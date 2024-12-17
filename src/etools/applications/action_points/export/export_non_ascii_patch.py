import re
import types

from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from tablib.formats._xlsx import XLSXFormat


def dset_sheet(cls, dataset, ws, freeze_panes=True, escape=False):
    """Completes given worksheet from given Dataset."""
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

    # Sanitize a single cell's value

    def sanitize_value(value):

        if value is None:
            return None
        if isinstance(value, str):
            # Remove illegal characters
            value = ILLEGAL_CHARACTERS_RE.sub('', value)
            # Replace NULL bytes or control characters
            value = re.sub(r'[\x00-\x1F\x7F]', '', value)
        return value

    _package = dataset._package(dicts=False)

    for i, sep in enumerate(dataset._separators):
        _offset = i
        _package.insert((sep[0] + _offset), (sep[1],))

    bold = Font(bold=True)
    wrap_text = Alignment(wrap_text=True)

    for i, row in enumerate(_package):
        row_number = i + 1
        for j, col in enumerate(row):
            col_idx = get_column_letter(j + 1)
            cell = ws[f'{col_idx}{row_number}']

            # Bold headers
            if (row_number == 1) and dataset.headers:
                cell.font = bold
                if freeze_panes:
                    # Export Freeze only after first Line
                    ws.freeze_panes = 'A2'

            # Bold separators
            elif len(row) < dataset.width:
                cell.font = bold

            # Wrap the rest
            else:
                if '\n' in str(col):
                    cell.alignment = wrap_text
            try:
                cell.value = sanitize_value(col)
            except Exception as e:
                # Log or print error for debugging
                print(f"Error processing cell at row {row_number}, col {j}: {e}")
                cell.value = str(col)  # Fallback to string

            if escape and cell.data_type == 'f' and cell.value.startswith('='):
                cell.value = cell.value.replace("=", "")


XLSXFormat.dset_sheet = types.MethodType(dset_sheet, XLSXFormat)

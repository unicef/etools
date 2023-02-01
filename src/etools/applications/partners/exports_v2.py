from copy import copy
from tempfile import NamedTemporaryFile

from django.utils.translation import ugettext as _

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework_csv import renderers as r

from etools.applications.core.renderers import FriendlyCSVRenderer
from etools.applications.core.util_scripts import currency_format
from etools.applications.partners.models import Intervention
from etools.applications.partners.utils import get_quarters_range


class PartnerOrganizationCSVRenderer(r.CSVRenderer):
    header = ['vendor_number', 'organization_full_name',
              'short_name', 'alternate_name', 'partner_type', 'shared_with', 'address',
              'phone_number', 'email_address', 'risk_rating', 'sea_risk_rating_nm', 'psea_assessment_date',
              'highest_risk_rating_type', 'highest_risk_rating_name', 'date_last_assessment_against_core_values',
              'actual_cash_transfer_for_cp', 'actual_cash_transfer_for_current_year', 'marked_for_deletion', 'blocked',
              'type_of_assessment', 'date_assessed', 'assessments', 'staff_members', 'url', 'planned_visits', ]

    labels = {
        'vendor_number': _('Vendor Number'),
        'organization_full_name': _('Organizations Full Name'),
        'short_name': _('Short Name'),
        'alternate_name': _('Alternate Name'),
        'partner_type': _('Partner Type'),
        'shared_with': _('Shared Partner'),
        'address': _('Address'),
        'phone_number': _('Phone Number'),
        'email_address': _('Email Address'),
        'risk_rating': _('HACT Risk Rating'),
        'date_last_assessment_against_core_values': _('Date Last Assessed Against Core Values'),
        'actual_cash_transfer_for_cp': _('Actual Cash Transfer for CP (USD)'),
        'actual_cash_transfer_for_current_year': _('Actual Cash Transfer for Current Year (USD)'),
        'marked_for_deletion': _('Marked for Deletion'),
        'blocked': _('Blocked'),
        'type_of_assessment': _('Assessment Type'),
        'date_assessed': _('Date Assessed'),
        'assessments': _('Assessment Type (Date Assessed)'),
        'staff_members': _('Staff Members'),
        'url': 'URL',
        'planned_visits': _("Planned Programmatic Visits"),
    }


class PartnerOrganizationHactCsvRenderer(FriendlyCSVRenderer):

    header = [
        'name',
        'vendor_number',
        'partner_type',
        'shared_with',
        'type_of_assessment',
        'net_ct_cy',
        'reported_cy',
        'total_ct_ytd',
        'rating',
        'flags.expiring_assessment_flag',
        'flags.approaching_threshold_flag',
        'hact_values.programmatic_visits.planned.q1',
        'hact_values.programmatic_visits.planned.q2',
        'hact_values.programmatic_visits.planned.q3',
        'hact_values.programmatic_visits.planned.q4',
        'hact_min_requirements.programmatic_visits',
        'hact_values.programmatic_visits.completed.q1',
        'hact_values.programmatic_visits.completed.q2',
        'hact_values.programmatic_visits.completed.q3',
        'hact_values.programmatic_visits.completed.q4',
        'planned_engagement.spot_check_planned_q1',
        'planned_engagement.spot_check_planned_q2',
        'planned_engagement.spot_check_planned_q3',
        'planned_engagement.spot_check_planned_q4',
        'hact_min_requirements.spot_checks',
        'planned_engagement.spot_check_follow_up',
        'hact_values.spot_checks.completed.q1',
        'hact_values.spot_checks.completed.q2',
        'hact_values.spot_checks.completed.q3',
        'hact_values.spot_checks.completed.q4',
        'hact_min_requirements.audits',
        'hact_values.audits.completed',
        'hact_values.outstanding_findings',
    ]

    labels = {
        'name': _('Implementing Partner'),
        'vendor_number': _('Vendor Number'),
        'partner_type': _('Partner Type'),
        'shared_with': _('Shared IP'),
        'type_of_assessment': _('Assessment Type'),
        'net_ct_cy': _('Cash Transfer 1 OCT - 30 SEP'),
        'reported_cy': _('Liquidations 1 OCT - 30 SEP'),
        'total_ct_ytd': _('Cash Transfers Jan - Dec'),
        'rating': _('Risk Rating'),
        'flags.expiring_assessment_flag': _('Expiring Threshold'),
        'flags.approaching_threshold_flag': _('Approach Threshold'),
        'hact_values.programmatic_visits.planned.q1': _('Programmatic Visits Planned Q1'),
        'hact_values.programmatic_visits.planned.q2': _('Q2'),
        'hact_values.programmatic_visits.planned.q3': _('Q3'),
        'hact_values.programmatic_visits.planned.q4': _('Q4'),
        'hact_min_requirements.programmatic_visits': _('Programmatic Visits M.R'),
        'hact_values.programmatic_visits.completed.q1': _('Programmatic Visits Completed Q1'),
        'hact_values.programmatic_visits.completed.q2': _('Q2'),
        'hact_values.programmatic_visits.completed.q3': _('Q3'),
        'hact_values.programmatic_visits.completed.q4': _('Q4'),
        'planned_engagement.spot_check_planned_q1': _('Spot Checks Planned Q1'),
        'planned_engagement.spot_check_planned_q2': _('Q2'),
        'planned_engagement.spot_check_planned_q3': _('Q3'),
        'planned_engagement.spot_check_planned_q4': _('Q4'),
        'hact_min_requirements.spot_checks': _('Spot Checks M.R'),
        'planned_engagement.spot_check_follow_up': _('Follow Up'),
        'hact_values.spot_checks.completed.q1': _('Spot Checks Completed Q1'),
        'hact_values.spot_checks.completed.q2': _('Q2'),
        'hact_values.spot_checks.completed.q3': _('Q3'),
        'hact_values.spot_checks.completed.q4': _('Q4'),
        'hact_min_requirements.audits': _('Audits M.R'),
        'hact_values.audits.completed': _('Audit Completed'),
        'hact_values.outstanding_findings': _('Audits Outstanding Findings'),
    }


class PartnerOrganizationDashboardCsvRenderer(FriendlyCSVRenderer):
    header = [
        'name',
        'sections',
        'locations',
        'action_points',
        'total_ct_cp',
        'total_ct_ytd',
        'outstanding_dct_amount_6_to_9_months_usd',
        'outstanding_dct_amount_more_than_9_months_usd',
        'days_last_pv',
        'alert_no_recent_pv',
        'alert_no_pv',
    ]

    labels = {
        'name': _('Implementing Partner'),
        'sections': _('Sections'),
        'locations': _('Locations'),
        'action_points': _('Action Points #'),
        'total_ct_cp': _('$ Cash in the Current CP Cycle'),
        'total_ct_ytd': _('$ Cash in the Current Year'),
        'days_last_pv': _('Days Since Last PV'),
        'alert_no_recent_pv': _('Alert: No Recent PV'),
        'alert_no_pv': _('Alert: No PV'),
        'outstanding_dct_amount_6_to_9_months_usd': _('Outstanding DCT Amount between 6 and 9 months'),
        'outstanding_dct_amount_more_than_9_months_usd': _('Outstanding DCT Amount more than 9 months'),
    }


class PartnerOrganizationSimpleHactCsvRenderer(FriendlyCSVRenderer):

    header = [
        'name',
        'vendor_number',
        'partner_type',
        'shared_with',
        'type_of_assessment',
        'total_ct_ytd',
        'rating',
        'flags.expiring_assessment_flag',
        'flags.approaching_threshold_flag',
        'hact_values.programmatic_visits.planned.total',
        'hact_min_requirements.programmatic_visits',
        'hact_values.programmatic_visits.completed.total',
        'planned_engagement.spot_check_required',
        'hact_values.spot_checks.completed.total',
        'hact_min_requirements.audits',
        'hact_values.audits.completed',
        'hact_values.outstanding_findings',
    ]

    labels = {
        'name': _('Implementing Partner'),
        'vendor_number': _('Vendor Number'),
        'partner_type': _('Partner Type'),
        'shared_with': _('Shared IP'),
        'total_ct_ytd': _('Cash Transfers Jan - Dec'),
        'type_of_assessment': _('Assessment Type'),
        'rating': _('Risk Rating'),
        'flags.expiring_assessment_flag': _('Expiring Threshold'),
        'flags.approaching_threshold_flag': _('Approach Threshold'),
        'hact_values.programmatic_visits.planned.total': _('Programmatic Visits Planned'),
        'hact_min_requirements.programmatic_visits': _('Programmatic Visits M.R'),
        'hact_values.programmatic_visits.completed.total': _('Programmatic Visits Completed'),
        'planned_engagement.spot_check_required': _('Spot Check Required'),
        'hact_values.spot_checks.completed.total': _('Spot Checks Completed'),
        'hact_min_requirements.audits': _('Audits M.R'),
        'hact_values.audits.completed': _('Audit Completed'),
        'hact_values.outstanding_findings': _('Audits Outstanding Findings'),
    }


class AgreementCSVRenderer(r.CSVRenderer):
    header = [
        "agreement_number",
        "status",
        "partner_name",
        "partner_number",
        "agreement_type",
        "start",
        "end",
        "partner_manager_name",
        "signed_by_partner_date",
        "signed_by_name",
        "signed_by_unicef_date",
        "staff_members",
        "amendments",
        "url",
        "special_conditions_pca",
        "terms_acknowledged_by",
    ]

    labels = {
        "agreement_number": _('Reference Number'),
        "status": _('Status'),
        "partner_name": _('Partner Name'),
        "partner_number": _("Vendor Number"),
        "agreement_type": _('Agreement Type'),
        "start": _('Start Date'),
        "end": _('End Date'),
        "partner_manager_name": _('Signed By Partner'),
        "signed_by_partner_date": _('Signed By Partner Date'),
        "signed_by_name": _('Signed By UNICEF'),
        "signed_by_unicef_date": _('Signed By UNICEF Date'),
        "staff_members": _('Partner Authorized Officer'),
        "amendments": _('Amendments'),
        "url": "URL",
        "special_conditions_pca": _("Special Conditions PCA"),
        "terms_acknowledged_by": _("Terms Acknowledged By"),
    }


class InterventionCSVRenderer(r.CSVRenderer):
    header = [
        "partner_name", "vendor_number", "status", "partner_type", "cso_type", "agreement_number", "country_programmes",
        "document_type", "number", "title", "start", "end", "offices", "sectors", "locations", "contingency_pd",
        "intervention_clusters", "unicef_focal_points", "partner_focal_points", "budget_currency", "cso_contribution",
        "unicef_budget", "unicef_supply", "total_planned_budget", "fr_numbers", "fr_currency", "fr_posting_date",
        "fr_amount", "fr_actual_amount", "fr_outstanding_amt", "planned_visits", "submission_date",
        "submission_date_prc", "review_date_prc", "partner_authorized_officer_signatory", "signed_by_partner_date",
        "unicef_signatory", "signed_by_unicef_date", "days_from_submission_to_signed", "days_from_review_to_signed",
        "amendment_sum", "last_amendment_date", "attachment_type", "total_attachments", "cp_outputs", "url",
        "cfei_number", "has_data_processing_agreement", "has_activities_involving_children",
        "has_special_conditions_for_construction",
    ]

    labels = {
        "partner_name": _("Partner"),
        "vendor_number": _("Vendor no."),
        "status": _("Status"),
        "partner_type": _("Partner Type"),
        "cso_type": _("CSO Type"),
        "agreement_number": _("Agreement"),
        "country_programmes": _("Country Programme"),
        "document_type": _("Document Type"),
        "number": _("Reference Number"),
        "title": _("Document Title"),
        "start": _("Start Date"),
        "end": _("End Date"),
        "offices": _("UNICEF Office"),
        "sectors": _("Sections"),
        "locations": _("Locations"),
        "contingency_pd": _("Contingency PD?"),
        "intervention_clusters": _("Cluster"),
        "unicef_focal_points": _("UNICEF Focal Points"),
        "partner_focal_points": _("CSO Authorized Officials"),
        "budget_currency": _("Budget Currency"),
        "cso_contribution": _("Total CSO Budget (USD)"),
        "unicef_budget": _("UNICEF Cash (USD)"),
        "unicef_supply": _("UNICEF Supply (USD)"),
        "total_planned_budget": _("Total PD/SPD Budget (USD)"),
        "fr_numbers": _("FR Number(s)"),
        "fr_currency": _("FR Currency"),
        "fr_posting_date": _("FR Posting Date"),
        "fr_amount": _("FR Amount"),
        "fr_actual_amount": _("FR Actual CT"),
        "fr_outstanding_amt": _("Outstanding DCT"),
        "planned_visits": _("Planned Programmatic Visits"),
        "spot_checks": _("Planned Spot Checks"),
        "audit": _("Planned Audits"),
        "submission_date": _("Document Submission Date by CSO"),
        "submission_date_prc": _("Submission Date to PRC"),
        "review_date_prc": _("Review Date by PRC"),
        "partner_authorized_officer_signatory": _("Signed by Partner"),
        "signed_by_partner_date": _("Signed by Partner Date"),
        "unicef_signatory": _("Signed by UNICEF"),
        "signed_by_unicef_date": _("Signed by UNICEF Date"),
        "days_from_submission_to_signed": _("Days from Submission to Signed"),
        "days_from_review_to_signed": _("Days from Review to Signed"),
        "amendment_sum": _("Total no. of amendments"),
        "last_amendment_date": _("Last amendment date"),
        "attachment_type": _("Attachment Type"),
        "total_attachments": _("# of attachments"),
        "cp_outputs": _("CP Outputs"),
        "url": "URL",
        "cfei_number": _("UNPP Number"),
        "has_data_processing_agreement": _("Data Processing Agreement"),
        "has_activities_involving_children": _("Activities involving children and young people"),
        "has_special_conditions_for_construction": _("Special Conditions for Construction Works by Implementing Partners"),
    }


class PartnershipDashCSVRenderer(r.CSVRenderer):
    header = [
        'partner_name',
        'partner_vendor_number',
        'number',
        'sections',
        'offices_names',
        'status',
        'start',
        'end',
        'budget_currency',
        'cso_contribution',
        'unicef_supplies',
        'unicef_cash',
        'fr_currency',
        'frs_total_frs_amt',
        'disbursement',
        'multi_curr_flag',
        'outstanding_dct',
        'frs_total_frs_amt_usd',
        'disbursement_usd',
        'outstanding_dct_usd',
        'disbursement_percent',
        'days_last_pv',
        'link',
    ]

    labels = {
        "partner_name": _("IP Name"),
        "partner_vendor_number": _("Vendor Number"),
        "number": _("PD/SPD Ref #"),
        "sections": _("Section"),
        "offices_names": _("Field Office"),
        "status": _("Status"),
        "start": _("Start Date"),
        "end": _("End Date"),
        "budget_currency": _("PD Currency"),
        "cso_contribution": _("CSO Contribution (PD Currency)"),
        "unicef_supplies": _("Total UNICEF Supplies (PD Currency)"),
        "unicef_cash": _("Total UNICEF Cash (PD Currency)"),
        "fr_currency": _("FR Currency"),
        "frs_total_frs_amt": _("FR Grand Total"),
        "disbursement": _("Actual Disbursements"),
        'multi_curr_flag': _("Multi-currency Transaction"),
        "outstanding_dct": _("Outstanding DCT"),
        "frs_total_frs_amt_usd": _("FR Grand Total (USD)"),
        "disbursement_usd": _("Actual Disbursement (USD)"),
        "outstanding_dct_usd": _("Outstanding DCT (USD)"),
        "disbursement_percent": _("Disbursement To Date (%)"),
        "days_last_pv": _("Days Since Last PV"),
        "link": _("Link")
    }


class InterventionLocationCSVRenderer(r.CSVRenderer):
    header = [   # This controls field order in the output
        'partner',
        'partner_vendor_number',
        'pd_ref_number',
        'partnership',
        'status',
        'location',
        'section',
        'cp_output',
        'start',
        'end',
        'focal_point',
        'hyperlink',
    ]
    labels = {
        'cp_output': _('CP output'),
        'end': _('End Date'),
        'focal_point': _('Name of UNICEF Focal Point'),
        'hyperlink': _('Hyperlink'),
        'location': _('Location'),
        'partner': _('Partner'),
        "partner_vendor_number": _("Vendor Number"),
        'partnership': _('Agreement'),
        'pd_ref_number': _('PD Ref Number'),
        'section': _('Section'),
        'start': _('Start Date'),
        'status': _('Status')
    }


class InterventionXLSRenderer:
    def __init__(self, intervention: Intervention):
        self.intervention = intervention
        self.color_black = '00000000'
        self.color_gray = '00C0C0C0'
        self.color_gray_dark = '00808080'
        self.color_blue = '00305496'
        self.color_blue_light = '0099CCFF'
        self.color_blue_pale = '00A6DBFE'
        self.color_blue_pale_light = '00CCEBFF'
        self.color_blue_border = '00B8CCE4'
        self.color_yellow = '00FFCC00'
        self.color_yellow_light = '00FDF0D2'

        self.fill_gray = ['fill', PatternFill(fill_type='solid', fgColor=self.color_gray)]
        self.fill_blue = ['fill', PatternFill(fill_type='solid', fgColor=self.color_blue)]
        self.fill_blue_light = ['fill', PatternFill(fill_type='solid', fgColor=self.color_blue_light)]
        self.fill_blue_pale = ['fill', PatternFill(fill_type='solid', fgColor=self.color_blue_pale)]
        self.fill_blue_pale_light = ['fill', PatternFill(fill_type='solid', fgColor=self.color_blue_pale_light)]
        self.fill_yellow = ['fill', PatternFill(fill_type='solid', fgColor=self.color_yellow)]
        self.fill_yellow_light = ['fill', PatternFill(fill_type='solid', fgColor=self.color_yellow_light)]

        self.font_default = Font(
            name='Calibri', size=11, bold=False, italic=False, vertAlign=None,
            underline='none', strike=False, color='FF000000'
        )
        self.font_white = ['font', copy(self.font_default)]
        self.font_white[1].color = '00FFFFFF'

        self.font_white_bold = ['font', copy(self.font_default)]
        self.font_white_bold[1].color = '00FFFFFF'
        self.font_white_bold[1].bold = True

        self.font_white_italic = ['font', copy(self.font_default)]
        self.font_white_italic[1].color = '00FFFFFF'
        self.font_white_italic[1].italic = True

        self.font_bold = ['font', copy(self.font_default)]
        self.font_bold[1].bold = True

        self.font_italic = ['font', copy(self.font_default)]
        self.font_italic[1].italic = True

        self.align_right = ["alignment", Alignment(horizontal='right')]
        self.align_center = ["alignment", Alignment(horizontal='center')]
        self.align_center_and_wrap = ["alignment", Alignment(horizontal='center', vertical='center', wrap_text=True)]

        self.border_blue_top_right = self.border(self.color_blue_border, 'thin', ['top', 'right'])
        self.border_blue_top_left_right = self.border(self.color_blue_border, 'thin', ['top', 'left', 'right'])
        self.border_blue_all = self.border(self.color_blue_border, 'thin', ['bottom', 'top', 'left', 'right'])

        self.border_black_top_right = self.border(self.color_black, 'thin', ['top', 'right'])
        self.border_black_top_left_right = self.border(self.color_black, 'thin', ['top', 'left', 'right'])
        self.border_black_all = self.border(self.color_black, 'thin', ['bottom', 'top', 'left', 'right'])

    def border(self, color, border_style, borders):
        return ['border', Border(**{side: Side(color=color, border_style=border_style) for side in borders})]

    def td_format(self, td_object):
        seconds = int(td_object.total_seconds())
        periods = [
            ('year', 60 * 60 * 24 * 365),
            ('month', 60 * 60 * 24 * 30),
            ('day', 60 * 60 * 24)
        ]

        strings = []
        for period_name, period_seconds in periods:
            if seconds > period_seconds:
                period_value, seconds = divmod(seconds, period_seconds)
                has_s = 's' if period_value > 1 else ''
                strings.append("%s %s%s" % (period_value, period_name, has_s))

        return ", ".join(strings)

    def apply_styles_to_cells(self, worksheet, start_row=1, start_column=1, end_row=1, end_column=1, styles=None):
        start_row, end_row = min(start_row, end_row), max(start_row, end_row)
        start_column, end_column = min(start_column, end_column), max(start_column, end_column)
        for i in range(start_row, end_row + 1):
            for j in range(start_column, end_column + 1):
                cell = worksheet[f'{get_column_letter(j)}{i}']
                for name, style in styles:
                    setattr(cell, name, style)

    def auto_format_cell_width(self, worksheet):
        for letter in range(1, worksheet.max_column):
            maximum_value = 0
            for cell in worksheet[get_column_letter(letter)]:
                val_to_check = len(str(cell.value))
                if val_to_check > maximum_value:
                    maximum_value = val_to_check
            worksheet.column_dimensions[get_column_letter(letter)].width = maximum_value + 1

    def apply_properties_to_columns(self, worksheet, start_column=1, column_count=1, properties=None):
        start_column = min(start_column, start_column + column_count - 1)
        end_column = max(start_column, start_column + column_count - 1)
        for i in range(start_column, end_column + 1):
            cell = worksheet.column_dimensions[get_column_letter(i)]
            for name, value in properties:
                setattr(cell, name, value)

    def sheet_header(self, worksheet, header_title):
        worksheet.append([header_title])
        worksheet.merge_cells(start_row=worksheet.max_row, start_column=1, end_row=worksheet.max_row, end_column=2)
        self.apply_styles_to_cells(worksheet, worksheet.max_row, 1, worksheet.max_row, 1, [self.font_bold])

        worksheet.append([
            _('Partner'),
            self.intervention.agreement.partner.name,
            _('Start date'), self.intervention.start.strftime('%d-%b-%y'),
            _('Currency %(currency)s') % {"currency": self.intervention.planned_budget.currency}
        ])
        worksheet.append([
            '',
            _('Vendor #: ') + self.intervention.agreement.partner.vendor_number,
            _('End date'), self.intervention.end.strftime('%d-%b-%y')
        ])
        worksheet.append([
            _('PD Reference'),
            self.intervention.number,
            _('Duration'),
            self.td_format(self.intervention.end - self.intervention.start)
        ])

        self.apply_styles_to_cells(worksheet, 2, 1, worksheet.max_row, 1, [self.font_bold])
        self.apply_styles_to_cells(worksheet, 2, 3, worksheet.max_row, 3, [self.font_bold])
        self.apply_styles_to_cells(worksheet, worksheet.max_row - 2, 4, worksheet.max_row, 4, [self.align_right])

        worksheet.append([''])

    def budget_summary(self, worksheet):
        budget = self.intervention.planned_budget
        unicef_contribution = budget.total_unicef_contribution_local()
        partner_contribution = budget.total_partner_contribution_local
        unicef_contribution_p = unicef_contribution / budget.total_local * 100 if budget.total_local else 0
        total_cash_p = budget.total_unicef_cash_local_wo_hq / unicef_contribution * 100 \
            if unicef_contribution else 0
        prog_effectivness_p = self.intervention.management_budgets.unicef_total / unicef_contribution * 100 \
            if unicef_contribution else 0
        supplies_p = budget.in_kind_amount_local / unicef_contribution * 100 \
            if unicef_contribution else 0
        partner_supplies_p = budget.partner_supply_local / partner_contribution * 100 \
            if partner_contribution else 0
        partner_total_cash = budget.partner_contribution_local / partner_contribution * 100 \
            if partner_contribution else 0
        partner_prog_effectivness_p = self.intervention.management_budgets.partner_total / partner_contribution * 100 \
            if partner_contribution else 0

        worksheet.append([_('Total PD Budget'), currency_format(budget.total_local), '%'])
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, 3, [self.fill_blue, self.font_white]
        )
        self.apply_styles_to_cells(worksheet, worksheet.max_row, 1, worksheet.max_row, 2, [self.font_white_bold])
        self.apply_styles_to_cells(worksheet, worksheet.max_row, 3, worksheet.max_row, 3, [self.font_white_italic])
        worksheet.append([
            _('UNICEF Contribution'),
            currency_format(unicef_contribution),
            '{:.2f}'.format(unicef_contribution_p)
        ])
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, 3, [self.font_bold, self.fill_blue_pale]
        )
        worksheet.append([
            _('Total Cash'),
            currency_format(budget.total_unicef_cash_local_wo_hq),
            '{:.2f}'.format(total_cash_p)
        ])
        self.apply_styles_to_cells(worksheet, worksheet.max_row, 1, worksheet.max_row, 3, [self.fill_yellow_light])
        worksheet.append([
            _('Prog effectiveness'),
            currency_format(self.intervention.management_budgets.unicef_total),
            '{:.2f}'.format(prog_effectivness_p)
        ])
        worksheet.append([
            _('HQ cost'),
            currency_format(budget.total_hq_cash_local),
            '{:.2f}'.format(self.intervention.hq_support_cost)
        ])
        worksheet.append(
            [_('Supplies in-kind'), currency_format(budget.in_kind_amount_local), '{:.2f}'.format(supplies_p)]
        )
        self.apply_styles_to_cells(worksheet, worksheet.max_row, 1, worksheet.max_row, 3, [self.fill_yellow_light])
        self.apply_styles_to_cells(worksheet, worksheet.max_row - 4, 3, worksheet.max_row, 3, [self.font_italic])

        worksheet.append([
            _('Partner Contribution'),
            currency_format(partner_contribution),
            '{:.2f}'.format(budget.partner_contribution_percent)
        ])
        self.apply_styles_to_cells(worksheet, worksheet.max_row, 1, worksheet.max_row, 3, [self.fill_blue_pale])
        self.apply_styles_to_cells(worksheet, worksheet.max_row, 1, worksheet.max_row, 2, [self.font_bold])
        self.apply_styles_to_cells(worksheet, worksheet.max_row, 3, worksheet.max_row, 3, [self.font_italic])
        worksheet.append([
            _('Total Cash'),
            currency_format(budget.partner_contribution_local),
            '{:.2f}'.format(partner_total_cash)
        ])
        self.apply_styles_to_cells(worksheet, worksheet.max_row, 1, worksheet.max_row, 3, [self.fill_yellow_light])
        worksheet.append([
            _('Prog effectiveness'),
            currency_format(self.intervention.management_budgets.partner_total),
            '{:.2f}'.format(partner_prog_effectivness_p)
        ])
        worksheet.append([
            _('Supplies in-kind'),
            currency_format(budget.partner_supply_local),
            '{:.2f}'.format(partner_supplies_p)
        ])
        self.apply_styles_to_cells(worksheet, worksheet.max_row, 1, worksheet.max_row, 3, [self.fill_yellow_light])
        self.apply_styles_to_cells(worksheet, worksheet.max_row - 1, 3, worksheet.max_row, 3, [self.font_italic])
        self.apply_styles_to_cells(worksheet, worksheet.max_row - 9, 2, worksheet.max_row, 3, [self.align_right])

        self.apply_styles_to_cells(
            worksheet, 7, 2, worksheet.max_row, 3, [self.border_black_top_right]
        )
        self.apply_styles_to_cells(
            worksheet, 7, 1, worksheet.max_row, 1, [self.border_black_top_left_right]
        )
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, 3, [self.border_black_all]
        )
        self.apply_styles_to_cells(
            worksheet, 6, 1, 6, 3, [self.border_blue_all]
        )

        worksheet.append([''])

        self.render_others_section(worksheet)
        self.auto_format_cell_width(worksheet)
        worksheet.column_dimensions['A'].width = 32

    def render_workplan_budget(self, worksheet):
        worksheet.append([
            '',
            _('PD Output/ PD Activity'),
            _('Total (%s)') % self.intervention.planned_budget.currency + '\n' + '(CSO + UNICEF)',
            _('CSO') + '\n' + _('contribution'),
            _('UNICEF') + '\n' + _('contribution'),
            _('PD Quarters')
        ])
        quarters = get_quarters_range(self.intervention.start, self.intervention.end)
        total_columns = 5 + len(quarters)
        worksheet.merge_cells(
            start_row=worksheet.max_row, start_column=6, end_row=worksheet.max_row, end_column=total_columns
        )
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, total_columns,
            [self.fill_blue, self.font_white_bold, self.border_blue_top_right, self.align_center]
        )
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, 1,
            [self.border_blue_top_left_right]
        )
        worksheet.append([
            _('Result Level'), '', '', '', ''
        ] + [f'Q{q.quarter}' for q in quarters])
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, total_columns,
            [self.fill_blue, self.font_white_bold, self.align_center, self.border_blue_top_right]
        )
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, 1,
            [self.border_blue_top_left_right]
        )

        for result_link in self.intervention.result_links.all():
            worksheet.append([
                _("CP Output ") + result_link.code + ":",
                result_link.cp_output.name if result_link.cp_output else ""
            ])
            self.apply_styles_to_cells(worksheet, worksheet.max_row, 1, worksheet.max_row, total_columns,
                                       [self.fill_blue_pale, self.font_bold, self.border_black_top_right])
            self.apply_styles_to_cells(worksheet, worksheet.max_row, 1, worksheet.max_row, 1,
                                       [self.border_black_top_left_right])
            worksheet.merge_cells(
                start_row=worksheet.max_row, start_column=1, end_row=worksheet.max_row, end_column=total_columns
            )
            for pd_output in result_link.ll_results.all():
                worksheet.append([
                    _("PD Output") + "\n" + pd_output.code + ":",
                    pd_output.name,
                    currency_format(pd_output.total()),
                    currency_format(pd_output.total_cso()),
                    currency_format(pd_output.total_unicef()),
                ])
                self.apply_styles_to_cells(
                    worksheet,
                    worksheet.max_row,
                    1,
                    worksheet.max_row,
                    total_columns,
                    [self.fill_blue_pale_light, self.border_black_top_right]
                )
                self.apply_styles_to_cells(
                    worksheet, worksheet.max_row, 1, worksheet.max_row, 1, [self.border_black_top_left_right]
                )
                self.apply_styles_to_cells(
                    worksheet, worksheet.max_row, 3, worksheet.max_row, total_columns, [self.font_bold]
                )
                worksheet.merge_cells(
                    start_row=worksheet.max_row, start_column=6, end_row=worksheet.max_row, end_column=total_columns
                )

                activities = pd_output.activities.all()
                for idx, activity in enumerate(activities):
                    time_frames = activity.time_frames.all()
                    worksheet.append([
                        activity.code + ' ' + str(idx) + ' ' + str(len(activities)),
                        activity.name,
                        currency_format(activity.total),
                        currency_format(activity.cso_cash),
                        currency_format(activity.unicef_cash),
                    ] +
                        ['x' if any(t.quarter == q.quarter for t in time_frames) else '' for q in quarters]
                    )

                    if idx < len(activities) - 1:
                        self.apply_styles_to_cells(
                            worksheet,
                            worksheet.max_row,
                            1,
                            worksheet.max_row,
                            total_columns,
                            [self.border_black_top_right]
                        )
                        self.apply_styles_to_cells(
                            worksheet, worksheet.max_row, 1, worksheet.max_row, 1, [self.border_black_top_left_right]
                        )
                    else:
                        self.apply_styles_to_cells(
                            worksheet,
                            worksheet.max_row,
                            1,
                            worksheet.max_row,
                            total_columns,
                            [self.border_black_all]
                        )

        worksheet.append([
            _('EEPM'), _('Effective and efficient programme management'),
            currency_format(self.intervention.management_budgets.total),
            currency_format(self.intervention.management_budgets.partner_total),
            currency_format(self.intervention.management_budgets.unicef_total),
        ])
        self.apply_styles_to_cells(
            worksheet,
            worksheet.max_row,
            1,
            worksheet.max_row,
            total_columns,
            [self.fill_blue_pale_light, self.border_black_top_right]
        )
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, 1, [self.border_black_top_left_right]
        )
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 3, worksheet.max_row, total_columns, [self.font_bold]
        )
        worksheet.merge_cells(
            start_row=worksheet.max_row, start_column=6, end_row=worksheet.max_row, end_column=total_columns
        )

        worksheet.append([
            _('EEPM.1'), _('In-country management & support'),
            currency_format(self.intervention.management_budgets.act1_unicef +
                            self.intervention.management_budgets.act1_partner),
            currency_format(self.intervention.management_budgets.act1_partner),
            currency_format(self.intervention.management_budgets.act1_unicef),
        ])
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, total_columns, [self.border_black_top_right]
        )
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, 1, [self.border_black_top_left_right]
        )
        worksheet.append([
            _('EEPM.2'), _('Operational costs'),
            currency_format(self.intervention.management_budgets.act2_unicef +
                            self.intervention.management_budgets.act2_partner),
            currency_format(self.intervention.management_budgets.act2_partner),
            currency_format(self.intervention.management_budgets.act2_unicef),
        ])
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, total_columns, [self.border_black_top_right]
        )
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, 1, [self.border_black_top_left_right]
        )
        worksheet.append([
            _('EEPM.3'), _('Planning, monitoring, evaluation, and communication'),
            currency_format(self.intervention.management_budgets.act3_unicef +
                            self.intervention.management_budgets.act3_partner),
            currency_format(self.intervention.management_budgets.act3_partner),
            currency_format(self.intervention.management_budgets.act3_unicef),
        ])
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, total_columns, [self.border_black_all]
        )
        worksheet.append([
            _('Subtotal for the programme costs'), '',
            currency_format(self.intervention.planned_budget.partner_contribution_local +
                            self.intervention.planned_budget.total_unicef_cash_local_wo_hq),
            currency_format(self.intervention.planned_budget.partner_contribution_local),
            currency_format(self.intervention.planned_budget.total_unicef_cash_local_wo_hq),
        ])
        self.apply_styles_to_cells(
            worksheet,
            worksheet.max_row,
            1,
            worksheet.max_row,
            total_columns,
            [self.fill_blue, self.font_white_bold, self.border_blue_top_right]
        )
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, 1, [self.border_blue_top_left_right]
        )
        worksheet.merge_cells(
            start_row=worksheet.max_row, start_column=6, end_row=worksheet.max_row, end_column=total_columns
        )
        worksheet.append([
            _('HQ Support/Capacity Building  ({%(cost)d}%% of UNICEF the cash component)')
            % {'cost': self.intervention.hq_support_cost},
            '', '', '',
            currency_format(self.intervention.planned_budget.total_hq_cash_local),
        ])
        self.apply_styles_to_cells(
            worksheet,
            worksheet.max_row,
            1,
            worksheet.max_row,
            total_columns,
            [self.fill_blue_pale_light, self.border_black_top_right, self.font_bold]
        )
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, 1, [self.border_black_top_left_right]
        )
        worksheet.merge_cells(
            start_row=worksheet.max_row, start_column=6, end_row=worksheet.max_row, end_column=total_columns
        )
        worksheet.append([
            _('Total PD budget cash'), '',
            currency_format(self.intervention.planned_budget.total_cash_local()),
            currency_format(self.intervention.planned_budget.partner_contribution_local),
            currency_format(self.intervention.planned_budget.unicef_cash_local),
        ])
        self.apply_styles_to_cells(
            worksheet,
            worksheet.max_row,
            1,
            worksheet.max_row,
            total_columns,
            [self.fill_blue, self.font_white_bold, self.border_blue_all]
        )
        worksheet.merge_cells(
            start_row=worksheet.max_row, start_column=6, end_row=worksheet.max_row, end_column=total_columns
        )

        self.apply_styles_to_cells(worksheet, 6, 1, 7, total_columns, [self.align_center_and_wrap])
        self.apply_styles_to_cells(worksheet, 7, 3, worksheet.max_row, 5, [self.align_right])
        self.auto_format_cell_width(worksheet)
        worksheet.column_dimensions['A'].width = 15
        self.apply_properties_to_columns(worksheet, 6, len(quarters), [['width', 5]])

    def render_detailed_workplan_budget(self, worksheet):
        quarters = get_quarters_range(self.intervention.start, self.intervention.end)
        worksheet.append([
            _('No.'),
            _('PD Output/ PD Activity / Item Description'),
            _('Unit Type'),
            _('Number of Units'),
            _('Price/Unit'),
            _('CSO') + '\n' + _('contribution'),
            _('UNICEF') + '\n' + _('contribution'),
            _('Total'),
            _('PD Quarters'),
        ] + [''] * (len(quarters) - 1) + [_('Other Notes')])

        total_columns = 9 + len(quarters)
        worksheet.merge_cells(
            start_row=worksheet.max_row, start_column=9, end_row=worksheet.max_row, end_column=total_columns - 1
        )
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, total_columns,
            [self.fill_blue, self.font_white_bold, self.align_center, self.border_blue_top_left_right]
        )
        worksheet.append(['', '', '', '', '', '', '', ''] + [f'Q{q.quarter}' for q in quarters] + [''])
        self.apply_properties_to_columns(worksheet, total_columns, total_columns, [['width', 50]])

        for i in range(1, 9):
            worksheet.merge_cells(
                start_row=worksheet.max_row - 1, start_column=i, end_row=worksheet.max_row, end_column=i
            )

        worksheet.merge_cells(
            start_row=worksheet.max_row - 1, start_column=total_columns,
            end_row=worksheet.max_row, end_column=total_columns
        )

        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, total_columns,
            [self.fill_blue, self.font_white_bold, self.align_center, self.border_blue_all]
        )
        for result_link in self.intervention.result_links.all():
            worksheet.append([
                result_link.code,
                _("CP Output ") + result_link.code + ": " + result_link.cp_output.name if result_link.cp_output else ""
            ])
            self.apply_styles_to_cells(
                worksheet, worksheet.max_row, 1, worksheet.max_row, total_columns, [self.fill_blue_pale, self.font_bold]
            )
            self.apply_styles_to_cells(
                worksheet, worksheet.max_row, 1, worksheet.max_row, 1, [self.border_black_top_left_right]
            )
            worksheet.merge_cells(
                start_row=worksheet.max_row,
                start_column=2,
                end_row=worksheet.max_row,
                end_column=total_columns - 1
            )
            for pd_output in result_link.ll_results.all():
                worksheet.append([
                    pd_output.code,
                    _("PD OUTPUT ") + pd_output.code + ": " + pd_output.name,
                    '',
                    '',
                    '',
                    currency_format(pd_output.total_cso()),
                    currency_format(pd_output.total_unicef()),
                    currency_format(pd_output.total())
                ])
                self.apply_styles_to_cells(
                    worksheet,
                    worksheet.max_row,
                    1,
                    worksheet.max_row,
                    total_columns,
                    [self.fill_blue_pale_light, self.font_bold]
                )
                worksheet.merge_cells(
                    start_row=worksheet.max_row, start_column=2, end_row=worksheet.max_row, end_column=5
                )

                for activity in pd_output.activities.all():
                    time_frames = activity.time_frames.all()
                    worksheet.append(
                        [
                            activity.code,
                            _("Activity") + ":" + activity.name,
                            '',
                            '',
                            '',
                            currency_format(activity.cso_cash),
                            currency_format(activity.unicef_cash),
                            currency_format(activity.total)
                        ] +
                        ['x' if any(t.quarter == q.quarter for t in time_frames) else '' for q in quarters] +
                        [activity.context_details]
                    )

                    self.apply_styles_to_cells(
                        worksheet,
                        worksheet.max_row,
                        1,
                        worksheet.max_row,
                        total_columns,
                        [self.fill_yellow_light, self.font_bold]
                    )
                    worksheet.merge_cells(
                        start_row=worksheet.max_row, start_column=2, end_row=worksheet.max_row, end_column=5
                    )

                    for item in activity.items.all():
                        worksheet.append([
                            item.code,
                            item.name,
                            item.unit,
                            item.no_units,
                            item.unit_price,
                            currency_format(item.cso_cash),
                            currency_format(item.unicef_cash),
                            currency_format(item.unicef_cash + item.cso_cash),
                        ])
                        self.apply_styles_to_cells(
                            worksheet, worksheet.max_row, 1, worksheet.max_row, total_columns, []
                        )

        worksheet.append([
            _('EEPM'), _('Effective and efficient programme management'),
            '', '', '',
            currency_format(self.intervention.management_budgets.total),
            currency_format(self.intervention.management_budgets.partner_total),
            currency_format(self.intervention.management_budgets.unicef_total),
        ])
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, total_columns, [self.fill_blue_pale_light]
        )
        worksheet.merge_cells(
            start_row=worksheet.max_row, start_column=2, end_row=worksheet.max_row, end_column=5
        )

        worksheet.append([
            _('EEPM.1'), _('Activity: In-country management & support'),
            '', '', '',
            currency_format(self.intervention.management_budgets.act1_partner),
            currency_format(self.intervention.management_budgets.act1_unicef),
            currency_format(self.intervention.management_budgets.act1_unicef +
                            self.intervention.management_budgets.act1_partner),
        ])
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, total_columns, [self.fill_yellow_light]
        )
        worksheet.merge_cells(
            start_row=worksheet.max_row, start_column=2, end_row=worksheet.max_row, end_column=5
        )

        for i, item in enumerate(self.intervention.management_budgets.items.filter(kind='in_country')):
            worksheet.append([
                _('EEPM.1.{%d}') % i + 1,
                item.name,
                '', '', '',
                currency_format(item.cso_cash),
                currency_format(item.unicef_cash),
                currency_format(item.unicef_cash + item.cso_cash),
            ])

        worksheet.append([
            _('EEPM.2'), _('Activity: Operational costs'),
            '', '', '',
            currency_format(self.intervention.management_budgets.act2_partner),
            currency_format(self.intervention.management_budgets.act2_unicef),
            currency_format(self.intervention.management_budgets.act2_unicef +
                            self.intervention.management_budgets.act2_partner),
        ])
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, total_columns, [self.fill_yellow_light]
        )
        worksheet.merge_cells(
            start_row=worksheet.max_row, start_column=2, end_row=worksheet.max_row, end_column=5
        )

        for i, item in enumerate(self.intervention.management_budgets.items.filter(kind='operational')):
            worksheet.append([
                _('EEPM.2.{%d}') % i + 1,
                item.name,
                '', '', '',
                currency_format(item.cso_cash),
                currency_format(item.unicef_cash),
                currency_format(item.unicef_cash + item.cso_cash),
            ])

        worksheet.append([
            _('EEPM.3'), _('Activity: Planning, monitoring, evaluation, and communication'),
            '', '', '',
            currency_format(self.intervention.management_budgets.act3_partner),
            currency_format(self.intervention.management_budgets.act3_unicef),
            currency_format(self.intervention.management_budgets.act3_unicef +
                            self.intervention.management_budgets.act3_partner),
        ])
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, total_columns, [self.fill_yellow_light]
        )
        worksheet.merge_cells(
            start_row=worksheet.max_row, start_column=2, end_row=worksheet.max_row, end_column=5
        )

        for i, item in enumerate(self.intervention.management_budgets.items.filter(kind='planning')):
            worksheet.append([
                _('EEPM.3.{%d}') % i + 1,
                item.name,
                '', '', '',
                currency_format(item.cso_cash),
                currency_format(item.unicef_cash),
                currency_format(item.unicef_cash + item.cso_cash),
            ])

        self.apply_styles_to_cells(
            worksheet, 8, 1, worksheet.max_row, total_columns, [self.border_black_top_right]
        )

        worksheet.append([
            _('Total Cost for all outputs'), '', '', '', '',
            currency_format(self.intervention.planned_budget.partner_contribution_local +
                            self.intervention.planned_budget.total_unicef_cash_local_wo_hq),
            currency_format(self.intervention.planned_budget.partner_contribution_local),
            currency_format(self.intervention.planned_budget.total_unicef_cash_local_wo_hq),
        ])
        self.apply_styles_to_cells(
            worksheet,
            worksheet.max_row,
            1,
            worksheet.max_row,
            total_columns,
            [self.fill_blue, self.font_white_bold, self.border_blue_all]
        )
        worksheet.merge_cells(
            start_row=worksheet.max_row, start_column=9, end_row=worksheet.max_row, end_column=total_columns - 1
        )
        worksheet.merge_cells(
            start_row=worksheet.max_row, start_column=1, end_row=worksheet.max_row, end_column=5
        )
        self.apply_styles_to_cells(worksheet, 6, 1, 6, total_columns, [self.align_center_and_wrap])
        self.apply_styles_to_cells(worksheet, 7, 3, worksheet.max_row, 8, [self.align_right])
        self.auto_format_cell_width(worksheet)
        worksheet.column_dimensions['A'].width = 15
        self.apply_properties_to_columns(worksheet, 9, len(quarters), [['width', 5]])
        self.apply_properties_to_columns(worksheet, total_columns, total_columns, [['width', 50]])

    def render_supply_plan(self, worksheet):
        worksheet.append([
            _('Item'), _('Number of Units'), _('Price/unit'), _('Total Price'),
            _('Provided By'), _('CP Output'), _('Other Mentions'), _('UNICEF Product Number')
        ])
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, 8, [self.fill_blue, self.font_white_bold]
        )

        supply_items_no = 0
        for supply_item in self.intervention.supply_items.all():
            worksheet.append([
                supply_item.title,
                currency_format(supply_item.unit_number),
                currency_format(supply_item.unit_price),
                currency_format(supply_item.total_price),
                supply_item.get_provided_by_display(),
                supply_item.result.cp_output.output_name if supply_item.result and supply_item.result.cp_output else "",
                supply_item.other_mentions,
                supply_item.unicef_product_number,
            ])
            supply_items_no += 1

        worksheet.append([''])

        self.apply_styles_to_cells(
            worksheet, worksheet.max_row - supply_items_no, 2, worksheet.max_row, 4, [self.align_right]
        )

        worksheet.append(
            [_('Total Value'), '', '', currency_format(self.intervention.planned_budget.in_kind_amount_local +
                                                       self.intervention.planned_budget.partner_supply_local)]
        )
        worksheet.append(
            [_('UNICEF Contribution'), '', '', currency_format(self.intervention.planned_budget.in_kind_amount_local)]
        )
        worksheet.append(
            [_('Partner Contribution'), '', '', currency_format(self.intervention.planned_budget.partner_supply_local)]
        )

        self.apply_styles_to_cells(
            worksheet, worksheet.max_row - 2, 1, worksheet.max_row, 4, [self.fill_blue, self.font_white]
        )
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row - 2, 4, worksheet.max_row, 4, [self.align_right]
        )
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row - 2, 1, worksheet.max_row - 2, 4, [self.font_white_bold]
        )

        self.auto_format_cell_width(worksheet)

    def render_others_section(self, worksheet):
        if self.intervention.ip_program_contribution:
            worksheet.append([_('Partner non-financial contribution:')])
            self.apply_styles_to_cells(
                worksheet, worksheet.max_row, 1, worksheet.max_row, 1, [self.fill_blue, self.font_white]
            )
            worksheet.append([self.intervention.ip_program_contribution])
            worksheet.append([''])

    def render(self):
        workbook = Workbook()

        if workbook.active:
            # remove default sheet
            workbook.remove(workbook.active)

        budget_summary_sheet = workbook.create_sheet(_('Budget Summary'))
        budget_summary_sheet.sheet_properties.tabColor = 'F4B183'
        activity_sheet = workbook.create_sheet(_('Activity Budget'))
        activity_sheet.sheet_properties.tabColor = '92D050'
        detailed_budget_sheet = workbook.create_sheet(_('Detailed Budget'))
        detailed_budget_sheet.sheet_properties.tabColor = 'FFD966'
        supply_cost_sheet = workbook.create_sheet(_('Supply Cost'))
        supply_cost_sheet.sheet_properties.tabColor = '00B0F0'

        self.sheet_header(budget_summary_sheet, _('Budget Summary'))
        self.budget_summary(budget_summary_sheet)

        self.sheet_header(activity_sheet, _('Workplan Budget'))
        self.render_workplan_budget(activity_sheet)

        self.sheet_header(detailed_budget_sheet, _('Detailed Workplan Budget'))
        self.render_detailed_workplan_budget(detailed_budget_sheet)

        self.sheet_header(supply_cost_sheet, _('Supply Contribution (Planned)'))
        self.render_supply_plan(supply_cost_sheet)

        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)
            data = tmp.read()

        return data

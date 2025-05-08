from copy import copy
from tempfile import NamedTemporaryFile

from django.utils.translation import gettext as _

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework_csv import renderers as r

from etools.applications.core.util_scripts import currency_format
from etools.applications.governments.models import GDD
from etools.applications.partners.utils import get_quarters_range


class GDDCSVRenderer(r.CSVRenderer):
    header = [
        "partner_name", "vendor_number", "status", "partner_type", "agreement_number", "country_programme",
        "number", "title", "start", "end", "offices", "sectors", "locations",
        "unicef_focal_points", "partner_focal_points", "budget_currency",
        "unicef_budget", "unicef_supply", "total_planned_budget", "fr_numbers", "fr_currency", "fr_posting_date",
        "fr_amount", "fr_actual_amount", "fr_outstanding_amt", "planned_visits", "submission_date",
        "submission_date_prc", "review_date_prc", "partner_authorized_officer_signatory", "signed_by_partner_date",
        "unicef_signatory", "signed_by_unicef_date", "days_from_submission_to_approved", "days_from_review_to_approved",
        "amendment_sum", "last_amendment_date", "attachment_type", "total_attachments", "cp_outputs", "url",
    ]

    labels = {
        "partner_name": _("Partner"),
        "vendor_number": _("Vendor no."),
        "status": _("Status"),
        "partner_type": _("Partner Type"),
        "agreement_number": _("Agreement"),
        "country_programme": _("Country Programme"),
        "number": _("Reference Number"),
        "title": _("Document Title"),
        "start": _("Start Date"),
        "end": _("End Date"),
        "offices": _("UNICEF Office"),
        "sectors": _("Contributing Sections"),
        "locations": _("Locations"),
        "unicef_focal_points": _("UNICEF Focal Points"),
        "partner_focal_points": _("Government Authorized Officials"),
        "budget_currency": _("Budget Currency"),
        "unicef_budget": _("UNICEF Cash (USD)"),
        "unicef_supply": _("UNICEF Supply (USD)"),
        "total_planned_budget": _("Total GPD Budget (USD)"),
        "fr_numbers": _("FR Number(s)"),
        "fr_currency": _("FR Currency"),
        "fr_posting_date": _("FR Posting Date"),
        "fr_amount": _("FR Amount"),
        "fr_actual_amount": _("FR Actual CT"),
        "fr_outstanding_amt": _("Outstanding DCT"),
        "planned_visits": _("Planned Programmatic Visits"),
        "spot_checks": _("Planned Spot Checks"),
        "audit": _("Planned Audits"),
        "submission_date": _("Document Submission Date by Government"),
        "submission_date_prc": _("Submission Date to PRC"),
        "review_date_prc": _("Review Date by PRC"),
        "partner_authorized_officer_signatory": _("Signed by Partner"),
        "signed_by_partner_date": _("Signed by Partner Date"),
        "unicef_signatory": _("Signed by UNICEF"),
        "signed_by_unicef_date": _("Signed by UNICEF Date"),
        "days_from_submission_to_approved": _("Days from Submission to Approved"),
        "days_from_review_to_approved": _("Days from Review to Approved"),
        "amendment_sum": _("Total no. of amendments"),
        "last_amendment_date": _("Last amendment date"),
        "attachment_type": _("Attachment Type"),
        "total_attachments": _("# of attachments"),
        "cp_outputs": _("CP Outputs"),
        "url": "URL",
    }


class GDDLocationCSVRenderer(r.CSVRenderer):
    header = [
        'partner', 'partner_vendor_number', 'pd_ref_number', 'partnership', 'status',
        'location', 'section', 'cp_output', 'start', 'end', 'focal_point', 'hyperlink',
    ]
    labels = {
        'cp_output': _('CP output'),
        'end': _('End Date'),
        'focal_point': _('Name of UNICEF Focal Point'),
        'hyperlink': _('Hyperlink'),
        'location': _('Location'),
        'partner': _('Partner'),
        "partner_vendor_number": _("Vendor Number"),
        'partnership': _('Agreement'),
        'pd_ref_number': _('PD Ref Number'),
        'section': _('Section'),
        'start': _('Start Date'),
        'status': _('Status')
    }


class GDDXLSRenderer:
    def __init__(self, gdd: GDD):
        self.gdd = gdd
        self.color_black = '00000000'
        self.color_gray = '00C0C0C0'
        self.color_gray_dark = '00808080'
        self.color_blue = '00305496'
        self.color_blue_light = '0099CCFF'
        self.color_blue_pale = '00A6DBFE'
        self.color_blue_pale_light = '00CCEBFF'
        self.color_blue_border = '00B8CCE4'
        self.color_yellow = '00FFCC00'
        self.color_yellow_light = '00FDF0D2'

        self.fill_gray = ['fill', PatternFill(fill_type='solid', fgColor=self.color_gray)]
        self.fill_blue = ['fill', PatternFill(fill_type='solid', fgColor=self.color_blue)]
        self.fill_blue_light = ['fill', PatternFill(fill_type='solid', fgColor=self.color_blue_light)]
        self.fill_blue_pale = ['fill', PatternFill(fill_type='solid', fgColor=self.color_blue_pale)]
        self.fill_blue_pale_light = ['fill', PatternFill(fill_type='solid', fgColor=self.color_blue_pale_light)]
        self.fill_yellow = ['fill', PatternFill(fill_type='solid', fgColor=self.color_yellow)]
        self.fill_yellow_light = ['fill', PatternFill(fill_type='solid', fgColor=self.color_yellow_light)]

        self.font_default = Font(
            name='Calibri', size=11, bold=False, italic=False, vertAlign=None,
            underline='none', strike=False, color='FF000000'
        )
        self.font_white = ['font', copy(self.font_default)]
        self.font_white[1].color = '00FFFFFF'

        self.font_white_bold = ['font', copy(self.font_default)]
        self.font_white_bold[1].color = '00FFFFFF'
        self.font_white_bold[1].bold = True

        self.font_white_italic = ['font', copy(self.font_default)]
        self.font_white_italic[1].color = '00FFFFFF'
        self.font_white_italic[1].italic = True

        self.font_bold = ['font', copy(self.font_default)]
        self.font_bold[1].bold = True

        self.font_italic = ['font', copy(self.font_default)]
        self.font_italic[1].italic = True

        self.align_right = ["alignment", Alignment(horizontal='right')]
        self.align_center = ["alignment", Alignment(horizontal='center')]
        self.align_center_and_wrap = ["alignment", Alignment(horizontal='center', vertical='center', wrap_text=True)]

        self.border_blue_top_right = self.border(self.color_blue_border, 'thin', ['top', 'right'])
        self.border_blue_top_left_right = self.border(self.color_blue_border, 'thin', ['top', 'left', 'right'])
        self.border_blue_all = self.border(self.color_blue_border, 'thin', ['bottom', 'top', 'left', 'right'])

        self.border_black_top_right = self.border(self.color_black, 'thin', ['top', 'right'])
        self.border_black_top_left_right = self.border(self.color_black, 'thin', ['top', 'left', 'right'])
        self.border_black_all = self.border(self.color_black, 'thin', ['bottom', 'top', 'left', 'right'])

    def border(self, color, border_style, borders):
        return ['border', Border(**{side: Side(color=color, border_style=border_style) for side in borders})]

    def td_format(self, td_object):
        seconds = int(td_object.total_seconds())
        periods = [
            ('year', 60 * 60 * 24 * 365),
            ('month', 60 * 60 * 24 * 30),
            ('day', 60 * 60 * 24)
        ]

        strings = []
        for period_name, period_seconds in periods:
            if seconds > period_seconds:
                period_value, seconds = divmod(seconds, period_seconds)
                has_s = 's' if period_value > 1 else ''
                strings.append("%s %s%s" % (period_value, period_name, has_s))

        return ", ".join(strings)

    def apply_styles_to_cells(self, worksheet, start_row=1, start_column=1, end_row=1, end_column=1, styles=None):
        start_row, end_row = min(start_row, end_row), max(start_row, end_row)
        start_column, end_column = min(start_column, end_column), max(start_column, end_column)
        for i in range(start_row, end_row + 1):
            for j in range(start_column, end_column + 1):
                cell = worksheet[f'{get_column_letter(j)}{i}']
                for name, style in styles:
                    setattr(cell, name, style)

    def auto_format_cell_width(self, worksheet):
        for letter in range(1, worksheet.max_column):
            maximum_value = 0
            for cell in worksheet[get_column_letter(letter)]:
                val_to_check = len(str(cell.value))
                if val_to_check > maximum_value:
                    maximum_value = val_to_check
            worksheet.column_dimensions[get_column_letter(letter)].width = maximum_value + 1

    def apply_properties_to_columns(self, worksheet, start_column=1, column_count=1, properties=None):
        start_column = min(start_column, start_column + column_count - 1)
        end_column = max(start_column, start_column + column_count - 1)
        for i in range(start_column, end_column + 1):
            cell = worksheet.column_dimensions[get_column_letter(i)]
            for name, value in properties:
                setattr(cell, name, value)

    def sheet_header(self, worksheet, header_title):
        worksheet.append([header_title])
        worksheet.merge_cells(start_row=worksheet.max_row, start_column=1, end_row=worksheet.max_row, end_column=2)
        self.apply_styles_to_cells(worksheet, worksheet.max_row, 1, worksheet.max_row, 1, [self.font_bold])

        worksheet.append([
            _('Partner'),
            self.gdd.partner.name,
            _('Start date'), self.gdd.start.strftime('%d-%b-%y') if self.gdd.start else '',
            _('Currency %(currency)s') % {"currency": self.gdd.planned_budget.currency}
        ])
        worksheet.append([
            '',
            _('Vendor #: ') + self.gdd.partner.vendor_number,
            _('End date'), self.gdd.end.strftime('%d-%b-%y') if self.gdd.end else ''
        ])
        worksheet.append([
            _('GPD Reference'),
            self.gdd.number,
            _('Duration'),
            self.td_format(self.gdd.end - self.gdd.start)
            if self.gdd.start and self.gdd.end else ''
        ])

        self.apply_styles_to_cells(worksheet, 2, 1, worksheet.max_row, 1, [self.font_bold])
        self.apply_styles_to_cells(worksheet, 2, 3, worksheet.max_row, 3, [self.font_bold])
        self.apply_styles_to_cells(worksheet, worksheet.max_row - 2, 4, worksheet.max_row, 4, [self.align_right])

        worksheet.append([''])

    def budget_summary(self, worksheet):
        budget = self.gdd.planned_budget
        unicef_contribution = budget.total_unicef_contribution_local()
        unicef_contribution_p = unicef_contribution / budget.total_local * 100 if budget.total_local else 0
        total_cash_p = budget.total_unicef_cash_local_wo_hq / unicef_contribution * 100 \
            if unicef_contribution else 0
        supplies_p = budget.in_kind_amount_local / unicef_contribution * 100 \
            if unicef_contribution else 0

        worksheet.append([_('Total GPD Budget'), currency_format(budget.total_local), '%'])
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, 3, [self.fill_blue, self.font_white]
        )
        self.apply_styles_to_cells(worksheet, worksheet.max_row, 1, worksheet.max_row, 2, [self.font_white_bold])
        self.apply_styles_to_cells(worksheet, worksheet.max_row, 3, worksheet.max_row, 3, [self.font_white_italic])
        worksheet.append([
            _('UNICEF Contribution'),
            currency_format(unicef_contribution),
            '{:.2f}'.format(unicef_contribution_p)
        ])
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, 3, [self.font_bold, self.fill_blue_pale]
        )
        worksheet.append([
            _('Total Cash'),
            currency_format(budget.total_unicef_cash_local_wo_hq),
            '{:.2f}'.format(total_cash_p)
        ])
        self.apply_styles_to_cells(worksheet, worksheet.max_row, 1, worksheet.max_row, 3, [self.fill_yellow_light])

        worksheet.append(
            [_('Supplies in-kind'), currency_format(budget.in_kind_amount_local), '{:.2f}'.format(supplies_p)]
        )
        self.apply_styles_to_cells(worksheet, worksheet.max_row, 1, worksheet.max_row, 3, [self.fill_yellow_light])
        self.apply_styles_to_cells(worksheet, worksheet.max_row - 4, 3, worksheet.max_row, 3, [self.font_italic])

        self.apply_styles_to_cells(worksheet, 7, 2, worksheet.max_row, 3, [self.border_black_top_right])
        self.apply_styles_to_cells(worksheet, 7, 1, worksheet.max_row, 1, [self.border_black_top_left_right])
        self.apply_styles_to_cells(worksheet, worksheet.max_row, 1, worksheet.max_row, 3, [self.border_black_all])
        self.apply_styles_to_cells(worksheet, 6, 1, 6, 3, [self.border_blue_all])

        worksheet.append([''])

        self.render_others_section(worksheet)
        self.auto_format_cell_width(worksheet)
        worksheet.column_dimensions['A'].width = 32

    def render_workplan_budget(self, worksheet):
        worksheet.append([
            '',
            _('GPD Output/ GPD Activity'),
            _('Total (%s)') % self.gdd.planned_budget.currency + '\n' + '(UNICEF)',
            _('UNICEF') + '\n' + _('contribution'),
            _('GPD Quarters')
        ])
        quarters = get_quarters_range(self.gdd.start, self.gdd.end)
        start_column = 6
        len_q = len(quarters) if quarters else 1
        total_columns = start_column + len_q
        worksheet.merge_cells(
            start_row=worksheet.max_row, start_column=start_column, end_row=worksheet.max_row, end_column=total_columns
        )
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, total_columns,
            [self.fill_blue, self.font_white_bold, self.border_blue_top_right, self.align_center]
        )
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, 1,
            [self.border_blue_top_left_right]
        )
        worksheet.append([
            _('Result Level'), '', '', '', ''
        ] + [f'Q{q.quarter}' for q in quarters] if quarters else [''])
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, total_columns,
            [self.fill_blue, self.font_white_bold, self.align_center, self.border_blue_top_right]
        )
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, 1,
            [self.border_blue_top_left_right]
        )

        for result_link in self.gdd.result_links.all():
            worksheet.append([
                _("CP Output ") + result_link.code + ":",
                result_link.cp_output.cp_output.name if result_link.cp_output else ""
            ])
            self.apply_styles_to_cells(worksheet, worksheet.max_row, 1, worksheet.max_row, total_columns,
                                       [self.fill_blue_pale, self.font_bold, self.border_black_top_right])
            self.apply_styles_to_cells(worksheet, worksheet.max_row, 1, worksheet.max_row, 1,
                                       [self.border_black_top_left_right])
            worksheet.merge_cells(
                start_row=worksheet.max_row, start_column=1, end_row=worksheet.max_row, end_column=total_columns
            )
            for ki in result_link.gdd_key_interventions.all():
                worksheet.append([
                    _("GPD Output") + "\n" + ki.code + ":",
                    ki.name,
                    currency_format(ki.total()),
                    currency_format(ki.total_unicef()),
                ])
                self.apply_styles_to_cells(
                    worksheet,
                    worksheet.max_row,
                    1,
                    worksheet.max_row,
                    total_columns,
                    [self.fill_blue_pale_light, self.border_black_top_right]
                )
                self.apply_styles_to_cells(
                    worksheet, worksheet.max_row, 1, worksheet.max_row, 1, [self.border_black_top_left_right]
                )
                self.apply_styles_to_cells(
                    worksheet, worksheet.max_row, 3, worksheet.max_row, total_columns, [self.font_bold]
                )
                worksheet.merge_cells(
                    start_row=worksheet.max_row, start_column=6, end_row=worksheet.max_row, end_column=total_columns
                )

                activities = ki.gdd_activities.all()
                for idx, activity in enumerate(activities):
                    time_frames = activity.time_frames.all()
                    worksheet.append([
                        activity.code + ' ' + str(idx) + ' ' + str(len(activities)),
                        activity.name if activity.is_active else f"({_('Inactive')}) {activity.name}",
                        currency_format(activity.total),
                        currency_format(activity.cso_cash),
                        currency_format(activity.unicef_cash),
                    ] +
                        ['x' if any(t.quarter == q.quarter for t in time_frames) else '' for q in quarters] if quarters else ['']
                    )

                    if idx < len(activities) - 1:
                        self.apply_styles_to_cells(
                            worksheet,
                            worksheet.max_row,
                            1,
                            worksheet.max_row,
                            total_columns,
                            [self.border_black_top_right]
                        )
                        self.apply_styles_to_cells(
                            worksheet, worksheet.max_row, 1, worksheet.max_row, 1, [self.border_black_top_left_right]
                        )
                    else:
                        self.apply_styles_to_cells(
                            worksheet,
                            worksheet.max_row,
                            1,
                            worksheet.max_row,
                            total_columns,
                            [self.border_black_all]
                        )

        # worksheet.append([
        #     _('EEPM'), _('Effective and efficient programme management'),
        #     currency_format(self.gdd.management_budgets.total),
        #     currency_format(self.gdd.management_budgets.partner_total),
        #     currency_format(self.gdd.management_budgets.unicef_total),
        # ])
        # self.apply_styles_to_cells(
        #     worksheet,
        #     worksheet.max_row,
        #     1,
        #     worksheet.max_row,
        #     total_columns,
        #     [self.fill_blue_pale_light, self.border_black_top_right]
        # )
        # self.apply_styles_to_cells(
        #     worksheet, worksheet.max_row, 1, worksheet.max_row, 1, [self.border_black_top_left_right]
        # )
        # self.apply_styles_to_cells(
        #     worksheet, worksheet.max_row, 3, worksheet.max_row, total_columns, [self.font_bold]
        # )
        # worksheet.merge_cells(
        #     start_row=worksheet.max_row, start_column=6, end_row=worksheet.max_row, end_column=total_columns
        # )
        #
        # worksheet.append([
        #     _('EEPM.1'), _('In-country management & support'),
        #     currency_format(self.gdd.management_budgets.act1_unicef +
        #                     self.gdd.management_budgets.act1_partner),
        #     currency_format(self.gdd.management_budgets.act1_partner),
        #     currency_format(self.gdd.management_budgets.act1_unicef),
        # ])
        # self.apply_styles_to_cells(
        #     worksheet, worksheet.max_row, 1, worksheet.max_row, total_columns, [self.border_black_top_right]
        # )
        # self.apply_styles_to_cells(
        #     worksheet, worksheet.max_row, 1, worksheet.max_row, 1, [self.border_black_top_left_right]
        # )
        # worksheet.append([
        #     _('EEPM.2'), _('Operational costs'),
        #     currency_format(self.gdd.management_budgets.act2_unicef +
        #                     self.gdd.management_budgets.act2_partner),
        #     currency_format(self.gdd.management_budgets.act2_partner),
        #     currency_format(self.gdd.management_budgets.act2_unicef),
        # ])
        # self.apply_styles_to_cells(
        #     worksheet, worksheet.max_row, 1, worksheet.max_row, total_columns, [self.border_black_top_right]
        # )
        # self.apply_styles_to_cells(
        #     worksheet, worksheet.max_row, 1, worksheet.max_row, 1, [self.border_black_top_left_right]
        # )
        # worksheet.append([
        #     _('EEPM.3'), _('Planning, monitoring, evaluation, and communication'),
        #     currency_format(self.gdd.management_budgets.act3_unicef +
        #                     self.gdd.management_budgets.act3_partner),
        #     currency_format(self.gdd.management_budgets.act3_partner),
        #     currency_format(self.gdd.management_budgets.act3_unicef),
        # ])
        # self.apply_styles_to_cells(
        #     worksheet, worksheet.max_row, 1, worksheet.max_row, total_columns, [self.border_black_all]
        # )
        worksheet.append([
            _('Subtotal for the programme costs'), '',
            currency_format(self.gdd.planned_budget.total_unicef_cash_local_wo_hq),
            currency_format(self.gdd.planned_budget.total_unicef_cash_local_wo_hq),
        ])
        self.apply_styles_to_cells(
            worksheet,
            worksheet.max_row,
            1,
            worksheet.max_row,
            total_columns,
            [self.fill_blue, self.font_white_bold, self.border_blue_top_right]
        )
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, 1, [self.border_blue_top_left_right]
        )
        worksheet.merge_cells(
            start_row=worksheet.max_row, start_column=6, end_row=worksheet.max_row, end_column=total_columns
        )
        # worksheet.append([
        #     _('Capacity Strengthening Costs ({%(cost)d}%% of UNICEF the cash component)')
        #     % {'cost': self.gdd.hq_support_cost},
        #     '', '', '',
        #     currency_format(self.gdd.planned_budget.total_hq_cash_local),
        # ])
        # self.apply_styles_to_cells(
        #     worksheet,
        #     worksheet.max_row,
        #     1,
        #     worksheet.max_row,
        #     total_columns,
        #     [self.fill_blue_pale_light, self.border_black_top_right, self.font_bold]
        # )
        # self.apply_styles_to_cells(
        #     worksheet, worksheet.max_row, 1, worksheet.max_row, 1, [self.border_black_top_left_right]
        # )
        # worksheet.merge_cells(
        #     start_row=worksheet.max_row, start_column=6, end_row=worksheet.max_row, end_column=total_columns
        # )
        worksheet.append([
            _('Total GPD budget cash'), '',
            currency_format(self.gdd.planned_budget.total_cash_local()),
            currency_format(self.gdd.planned_budget.unicef_cash_local),
        ])
        self.apply_styles_to_cells(
            worksheet,
            worksheet.max_row,
            1,
            worksheet.max_row,
            total_columns,
            [self.fill_blue, self.font_white_bold, self.border_blue_all]
        )
        worksheet.merge_cells(
            start_row=worksheet.max_row, start_column=6, end_row=worksheet.max_row, end_column=total_columns
        )

        self.apply_styles_to_cells(worksheet, 6, 1, 7, total_columns, [self.align_center_and_wrap])
        self.apply_styles_to_cells(worksheet, 7, 3, worksheet.max_row, 5, [self.align_right])
        self.auto_format_cell_width(worksheet)
        worksheet.column_dimensions['A'].width = 15
        self.apply_properties_to_columns(worksheet, 6, len_q, [['width', 5]])

    def render_detailed_workplan_budget(self, worksheet):
        quarters = get_quarters_range(self.gdd.start, self.gdd.end)
        worksheet.append([
            _('No.'),
            _('GPD Output/ GPD Activity / Item Description'),
            _('Unit Type'),
            _('Number of Units'),
            _('Price/Unit'),
            _('UNICEF') + '\n' + _('contribution'),
            _('Total'),
            _('GPD Quarters'),
        ] + [''] * (len(quarters) - 1) + [_('Other Notes')])

        start_column = 9
        total_columns = start_column + len(quarters)
        worksheet.merge_cells(
            start_row=worksheet.max_row, start_column=start_column, end_row=worksheet.max_row, end_column=total_columns
        )
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, total_columns,
            [self.fill_blue, self.font_white_bold, self.align_center, self.border_blue_top_left_right]
        )
        worksheet.append(['', '', '', '', '', '', '', ''] + [f'Q{q.quarter}' for q in quarters] if quarters else [''] + [''])
        self.apply_properties_to_columns(worksheet, total_columns, total_columns, [['width', 50]])

        for i in range(1, 9):
            worksheet.merge_cells(
                start_row=worksheet.max_row - 1, start_column=i, end_row=worksheet.max_row, end_column=i
            )

        worksheet.merge_cells(
            start_row=worksheet.max_row - 1, start_column=total_columns,
            end_row=worksheet.max_row, end_column=total_columns
        )

        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, total_columns,
            [self.fill_blue, self.font_white_bold, self.align_center, self.border_blue_all]
        )
        for result_link in self.gdd.result_links.all():
            worksheet.append([
                result_link.code,
                _("CP Output ") + result_link.code + ": " + result_link.cp_output.cp_output.name if result_link.cp_output else ""
            ])
            self.apply_styles_to_cells(
                worksheet, worksheet.max_row, 1, worksheet.max_row, total_columns, [self.fill_blue_pale, self.font_bold]
            )
            self.apply_styles_to_cells(
                worksheet, worksheet.max_row, 1, worksheet.max_row, 1, [self.border_black_top_left_right]
            )
            worksheet.merge_cells(
                start_row=worksheet.max_row,
                start_column=2,
                end_row=worksheet.max_row,
                end_column=total_columns - 1
            )
            for ki in result_link.gdd_key_interventions.all():
                worksheet.append([
                    ki.code,
                    _("GPD OUTPUT ") + ki.code + ": " + ki.name,
                    '',
                    '',
                    '',
                    currency_format(ki.total_unicef()),
                    currency_format(ki.total())
                ])
                self.apply_styles_to_cells(
                    worksheet,
                    worksheet.max_row,
                    1,
                    worksheet.max_row,
                    total_columns,
                    [self.fill_blue_pale_light, self.font_bold]
                )
                worksheet.merge_cells(
                    start_row=worksheet.max_row, start_column=2, end_row=worksheet.max_row, end_column=5
                )

                for activity in ki.gdd_activities.all():
                    time_frames = activity.time_frames.all()
                    title = _("Activity") + ":" + activity.name
                    worksheet.append(
                        [
                            activity.code,
                            title if activity.is_active else f"({_('Inactive')}) {title}",
                            '',
                            '',
                            '',
                            currency_format(activity.unicef_cash),
                            currency_format(activity.total)
                        ] +
                        ['x' if any(t.quarter == q.quarter for t in time_frames) else '' for q in quarters] if quarters else [''] +
                        [activity.context_details]
                    )

                    self.apply_styles_to_cells(
                        worksheet,
                        worksheet.max_row,
                        1,
                        worksheet.max_row,
                        total_columns,
                        [self.fill_yellow_light, self.font_bold]
                    )
                    worksheet.merge_cells(
                        start_row=worksheet.max_row, start_column=2, end_row=worksheet.max_row, end_column=5
                    )

                    for item in activity.items.all():
                        worksheet.append([
                            item.code,
                            item.name,
                            item.unit,
                            item.no_units,
                            item.unit_price,
                            currency_format(item.unicef_cash),
                        ])
                        self.apply_styles_to_cells(
                            worksheet, worksheet.max_row, 1, worksheet.max_row, total_columns, []
                        )

        worksheet.append([
            _('Total Cost for all outputs'), '', '', '', '',
            currency_format(self.gdd.planned_budget.partner_contribution_local),
            currency_format(self.gdd.planned_budget.total_unicef_cash_local_wo_hq),
            currency_format(self.gdd.planned_budget.partner_contribution_local +
                            self.gdd.planned_budget.total_unicef_cash_local_wo_hq),
        ])
        self.apply_styles_to_cells(
            worksheet,
            worksheet.max_row,
            1,
            worksheet.max_row,
            total_columns,
            [self.fill_blue, self.font_white_bold, self.border_blue_all]
        )
        worksheet.merge_cells(
            start_row=worksheet.max_row, start_column=start_column, end_row=worksheet.max_row, end_column=total_columns
        )
        worksheet.merge_cells(
            start_row=worksheet.max_row, start_column=1, end_row=worksheet.max_row, end_column=5
        )
        self.apply_styles_to_cells(worksheet, 6, 1, 6, total_columns, [self.align_center_and_wrap])
        self.apply_styles_to_cells(worksheet, 7, 3, worksheet.max_row, 8, [self.align_right])
        self.auto_format_cell_width(worksheet)
        worksheet.column_dimensions['A'].width = 15
        self.apply_properties_to_columns(worksheet, 9, len(quarters) if quarters else 1, [['width', 5]])
        self.apply_properties_to_columns(worksheet, total_columns, total_columns, [['width', 50]])

    def render_supply_plan(self, worksheet):
        worksheet.append([
            _('Item'), _('Number of Units'), _('Price/unit'), _('Total Price'),
            _('Provided By'), _('CP Output'), _('Other Mentions'), _('UNICEF Product Number')
        ])
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row, 1, worksheet.max_row, 8, [self.fill_blue, self.font_white_bold]
        )

        supply_items_no = 0
        for supply_item in self.gdd.supply_items.all():
            worksheet.append([
                supply_item.title,
                currency_format(supply_item.unit_number),
                currency_format(supply_item.unit_price),
                currency_format(supply_item.total_price),
                supply_item.get_provided_by_display(),
                supply_item.result.cp_output.cp_output.name if supply_item.result and supply_item.result.cp_output else "",
                supply_item.other_mentions,
                supply_item.unicef_product_number,
            ])
            supply_items_no += 1

        worksheet.append([''])

        self.apply_styles_to_cells(
            worksheet, worksheet.max_row - supply_items_no, 2, worksheet.max_row, 4, [self.align_right]
        )

        worksheet.append(
            [_('Total Value'), '', '', currency_format(self.gdd.planned_budget.in_kind_amount_local +
                                                       self.gdd.planned_budget.partner_supply_local)]
        )
        worksheet.append(
            [_('UNICEF Contribution'), '', '', currency_format(self.gdd.planned_budget.in_kind_amount_local)]
        )

        self.apply_styles_to_cells(
            worksheet, worksheet.max_row - 2, 1, worksheet.max_row, 4, [self.fill_blue, self.font_white]
        )
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row - 2, 4, worksheet.max_row, 4, [self.align_right]
        )
        self.apply_styles_to_cells(
            worksheet, worksheet.max_row - 2, 1, worksheet.max_row - 2, 4, [self.font_white_bold]
        )

        self.auto_format_cell_width(worksheet)

    def render_others_section(self, worksheet):
        if self.gdd.ip_program_contribution:
            worksheet.append([_('Partner non-financial contribution:')])
            self.apply_styles_to_cells(
                worksheet, worksheet.max_row, 1, worksheet.max_row, 1, [self.fill_blue, self.font_white]
            )
            worksheet.append([self.gdd.ip_program_contribution])
            worksheet.append([''])

    def render(self):
        workbook = Workbook()

        if workbook.active:
            # remove default sheet
            workbook.remove(workbook.active)

        budget_summary_sheet = workbook.create_sheet(_('Budget Summary'))
        budget_summary_sheet.sheet_properties.tabColor = 'F4B183'
        activity_sheet = workbook.create_sheet(_('Activity Budget'))
        activity_sheet.sheet_properties.tabColor = '92D050'
        detailed_budget_sheet = workbook.create_sheet(_('Detailed Budget'))
        detailed_budget_sheet.sheet_properties.tabColor = 'FFD966'
        supply_cost_sheet = workbook.create_sheet(_('Supply Cost'))
        supply_cost_sheet.sheet_properties.tabColor = '00B0F0'

        self.sheet_header(budget_summary_sheet, _('Budget Summary'))
        self.budget_summary(budget_summary_sheet)

        self.sheet_header(activity_sheet, _('Workplan Budget'))
        self.render_workplan_budget(activity_sheet)

        self.sheet_header(detailed_budget_sheet, _('Detailed Workplan Budget'))
        self.render_detailed_workplan_budget(detailed_budget_sheet)

        self.sheet_header(supply_cost_sheet, _('Supply Contribution (Planned)'))
        self.render_supply_plan(supply_cost_sheet)

        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)
            data = tmp.read()

        return data

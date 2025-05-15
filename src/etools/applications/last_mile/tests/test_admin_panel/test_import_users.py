import io
import json

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.http import HttpResponse

import openpyxl
from rest_framework import status
from rest_framework.reverse import reverse

from etools.applications.core.tests.cases import BaseTenantTestCase
from etools.applications.last_mile.admin_panel.constants import *  # NOQA
from etools.applications.last_mile.admin_panel.constants import ADMIN_PANEL_APP_NAME, USER_ADMIN_PANEL
from etools.applications.last_mile.models import Profile as LastMileProfile
from etools.applications.last_mile.tests.factories import PointOfInterestFactory
from etools.applications.organizations.tests.factories import OrganizationFactory
from etools.applications.partners.tests.factories import PartnerFactory
from etools.applications.users.models import Realm
from etools.applications.users.tests.factories import UserPermissionFactory

User = get_user_model()


class TestUserAdminViewSetImport(BaseTenantTestCase):

    @classmethod
    def setUpTestData(cls):
        cls.admin_user = UserPermissionFactory(
            realms__data=['LMSM Admin Panel', 'IP LM Editor'],
            username="importer_admin",
            email="importer_admin@example.com",
            is_staff=True,
            perms=[USER_ADMIN_PANEL_PERMISSION]
        )
        cls.ip_lm_editor_group = Group.objects.get(name="IP LM Editor")
        cls.org1 = OrganizationFactory(name="Import Org 1", vendor_number="VN001")
        cls.partner1 = PartnerFactory(organization=cls.org1)
        cls.poi1_org1 = PointOfInterestFactory(name="POI 1 Org 1", p_code="PCODE001", partner_organizations=[cls.partner1])
        cls.poi2_org1 = PointOfInterestFactory(name="POI 2 Org 1", p_code="PCODE002", partner_organizations=[cls.partner1])

        cls.org2 = OrganizationFactory(name="Import Org 2", vendor_number="VN002")
        cls.partner2 = PartnerFactory(organization=cls.org2)
        cls.poi1_org2 = PointOfInterestFactory(name="POI 1 Org 2", p_code="PCODE003", partner_organizations=[cls.partner2])

        cls.unlinked_poi = PointOfInterestFactory(name="Unlinked POI", p_code="PCODE999")

        cls.import_url = reverse(f'{ADMIN_PANEL_APP_NAME}:{USER_ADMIN_PANEL}--import-file')

    def _create_xlsx_file(self, data_rows, headers=None):
        if headers is None:
            headers = ["ip_number", "first_name", "last_name", "email", "point_of_interests"]

        stream = io.BytesIO()
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        sheet.append(headers)
        sheet.append(["Help text / Description row (ignored by importer)"])

        for row_data in data_rows:
            sheet.append(row_data)

        workbook.save(stream)
        stream.seek(0)
        return stream

    def test_import_users_successful(self):
        self.assertEqual(User.objects.count(), 1)
        data_rows = [
            ["VN001", "Alice", "Smith", "alice.smith@example.com", json.dumps(["PCODE001", "PCODE002"])],
            ["VN002", "Bob", "Johnson", "bob.johnson@example.com", json.dumps(["PCODE003"])],
            ["VN001", "Charlie", "Brown", "charlie.brown@example.com", json.dumps([])],
        ]
        xlsx_file = self._create_xlsx_file(data_rows)
        xlsx_file.name = "success_import.xlsx"

        response = self.forced_auth_req(
            'post',
            self.import_url,
            user=self.admin_user,
            data={'file': xlsx_file},
            request_format='multipart'
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data, {"valid": True})

        self.assertEqual(User.objects.count(), 1 + 3)

        alice = User.objects.get(email="alice.smith@example.com")
        self.assertEqual(alice.first_name, "Alice")
        self.assertEqual(alice.last_name, "Smith")
        self.assertEqual(alice.username, "alice.smith@example.com")
        self.assertFalse(alice.is_active)
        self.assertTrue(alice.check_password("test_pass"))
        self.assertEqual(alice.profile.organization, self.org1)
        self.assertEqual(alice.profile.job_title, "")
        self.assertEqual(alice.profile.phone_number, "")
        self.assertTrue(Realm.objects.filter(user=alice, group=self.ip_lm_editor_group, organization=self.org1).exists())
        lmp = LastMileProfile.objects.get(user=alice)
        self.assertEqual(lmp.created_by, self.admin_user)

        bob = User.objects.get(email="bob.johnson@example.com")
        self.assertEqual(bob.profile.organization, self.org2)
        self.assertTrue(Realm.objects.filter(user=bob, group=self.ip_lm_editor_group, organization=self.org2).exists())
        self.assertEqual(set(p.p_code for p in bob.profile.organization.partner.points_of_interest.all()), {"PCODE003"})

        charlie = User.objects.get(email="charlie.brown@example.com")
        self.assertEqual(charlie.profile.organization, self.org1)

        self.assertEqual(self.partner1.points_of_interest.count(), 0)

    def test_import_users_no_file_provided(self):
        response = self.forced_auth_req(
            'post',
            self.import_url,
            user=self.admin_user,
            data={},
            request_format='multipart'
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('file', response.data)

    def test_import_users_with_some_invalid_rows(self):
        initial_user_count = User.objects.count()
        data_rows = [
            ["VN001", "Valid", "User1", "valid.user1@example.com", json.dumps(["PCODE001"])],
            ["VN999", "InvalidOrg", "User2", "invalid.org@example.com", json.dumps([])],
            ["VN002", "ValidToo", "User3", "bademail", json.dumps([])],
            ["VN002", "David", "Copper", "david.copper@example.com", "not_json_pois"],
            ["VN001", "Eve", "Online", "eve.online@example.com", json.dumps(["PCODE_NONEXIST"])],
        ]
        xlsx_file = self._create_xlsx_file(data_rows)
        xlsx_file.name = "partial_fail_import.xlsx"

        response = self.forced_auth_req(
            'post',
            self.import_url,
            user=self.admin_user,
            data={'file': xlsx_file},
            request_format='multipart'
        )

        self.assertIsInstance(response, HttpResponse)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response['Content-Disposition'].startswith('attachment; filename="checked_partial_fail_import.xlsx"'))
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        self.assertEqual(User.objects.count(), initial_user_count + 1)
        self.assertTrue(User.objects.filter(email="valid.user1@example.com").exists())
        self.assertFalse(User.objects.filter(email="invalid.org@example.com").exists())
        self.assertFalse(User.objects.filter(email__iexact="bademail").exists())
        self.assertFalse(User.objects.filter(email="david.copper@example.com").exists())
        self.assertFalse(User.objects.filter(email="eve.online@example.com").exists())

        returned_file_content = io.BytesIO(response.content)
        workbook = openpyxl.load_workbook(returned_file_content)
        sheet = workbook.active
        errors_col_idx = sheet.max_column

        self.assertEqual(sheet.cell(row=3, column=errors_col_idx).value, "Success")
        self.assertIn("Organization not found by vendor number", sheet.cell(row=4, column=errors_col_idx).value)
        self.assertIn("Enter a valid email address.", sheet.cell(row=5, column=errors_col_idx).value)
        self.assertIn("Invalid point of interest format", sheet.cell(row=6, column=errors_col_idx).value)
        self.assertIn("does not exist", sheet.cell(row=7, column=errors_col_idx).value)

    def test_import_users_all_rows_invalid_returns_error_file(self):
        initial_user_count = User.objects.count()
        data_rows = [
            ["VN999", "Invalid1", "User", "invalid1@example.com", json.dumps([])],
            ["VN888", "Invalid2", "User", "invalid2@example.com", json.dumps([])],
        ]
        xlsx_file = self._create_xlsx_file(data_rows)
        xlsx_file.name = "all_fail_import.xlsx"

        response = self.forced_auth_req(
            'post',
            self.import_url,
            user=self.admin_user,
            data={'file': xlsx_file},
            request_format='multipart'
        )
        self.assertIsInstance(response, HttpResponse)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response['Content-Disposition'].startswith('attachment; filename="checked_all_fail_import.xlsx"'))
        self.assertEqual(User.objects.count(), initial_user_count)

    def test_import_users_duplicate_email_in_db_fails_creation_in_file(self):
        User.objects.create_user(username="existing.user@example.com", email="existing.user@example.com", password="password")
        initial_user_count = User.objects.count()

        data_rows = [
            ["VN001", "New", "User", "new.user@example.com", json.dumps([])],
            ["VN002", "Existing", "User", "existing.user@example.com", json.dumps([])],
        ]
        xlsx_file = self._create_xlsx_file(data_rows)
        xlsx_file.name = "duplicate_email_import.xlsx"

        response = self.forced_auth_req(
            'post',
            self.import_url,
            user=self.admin_user,
            data={'file': xlsx_file},
            request_format='multipart'
        )

        self.assertIsInstance(response, HttpResponse)
        self.assertEqual(response.status_code, 200)

        self.assertEqual(User.objects.count(), initial_user_count + 1)
        self.assertTrue(User.objects.filter(email="new.user@example.com").exists())
        self.assertTrue(User.objects.filter(email="existing.user@example.com").exists())

        returned_file_content = io.BytesIO(response.content)
        workbook = openpyxl.load_workbook(returned_file_content)
        sheet = workbook.active
        errors_col_idx = sheet.max_column
        self.assertEqual(sheet.cell(row=3, column=errors_col_idx).value, "Success")
        self.assertTrue(
            "UNIQUE constraint failed" in sheet.cell(row=4, column=errors_col_idx).value.lower() or
            "already exists" in sheet.cell(row=4, column=errors_col_idx).value.lower()
        )

    def test_import_users_empty_point_of_interests_string(self):
        data_rows = [
            ["VN001", "EmptyPOI", "String", "emptypoi.string@example.com", ""],
        ]
        xlsx_file = self._create_xlsx_file(data_rows)
        xlsx_file.name = "empty_poi_string.xlsx"

        response = self.forced_auth_req(
            'post',
            self.import_url,
            user=self.admin_user,
            data={'file': xlsx_file},
            request_format='multipart'
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data, {"valid": True})

        user = User.objects.get(email="emptypoi.string@example.com")
        self.assertEqual(user.profile.organization.partner.points_of_interest.count(), 0)

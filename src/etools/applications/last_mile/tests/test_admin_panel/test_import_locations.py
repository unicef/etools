import io
import json
from unittest.mock import patch

from django.http import HttpResponse

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from rest_framework import status
from rest_framework.reverse import reverse

from etools.applications.core.tests.cases import BaseTenantTestCase
from etools.applications.last_mile.admin_panel.constants import *  # NOQA
from etools.applications.last_mile.admin_panel.constants import ADMIN_PANEL_APP_NAME
from etools.applications.last_mile.models import PointOfInterest
from etools.applications.last_mile.tests.factories import PointOfInterestFactory, PointOfInterestTypeFactory
from etools.applications.organizations.tests.factories import OrganizationFactory
from etools.applications.partners.tests.factories import PartnerFactory
from etools.applications.users.tests.factories import UserPermissionFactory


class TestLocationAdminViewSetImport(BaseTenantTestCase):

    @classmethod
    def setUpTestData(cls):
        cls.admin_user = UserPermissionFactory(
            realms__data=['LMSM Admin Panel'],
            username="lmsm_importer_admin",
            email="lmsm_importer_admin@example.com",
            is_staff=True,
            perms=[USER_ADMIN_PANEL_PERMISSION, LOCATIONS_ADMIN_PANEL_PERMISSION]
        )

        cls.poi_type_health = PointOfInterestTypeFactory(name="Health Facility")
        cls.poi_type_school = PointOfInterestTypeFactory(name="School")
        cls.poi_type_other = PointOfInterestTypeFactory(name="Other")

        cls.org1 = OrganizationFactory(name="IP Org 1", vendor_number="IP_VN_001")
        cls.partner_org1 = PartnerFactory(organization=cls.org1)

        cls.org2 = OrganizationFactory(name="IP Org 2", vendor_number="IP_VN_002")
        cls.partner_org2 = PartnerFactory(organization=cls.org2)

        cls.existing_poi = PointOfInterestFactory(
            name="Existing General Hospital",
            p_code="PCODE_EXISTING",
            poi_type=cls.poi_type_health
        )

        cls.import_url = reverse(f'{ADMIN_PANEL_APP_NAME}:{LOCATIONS_ADMIN_PANEL}--import-file')

    def _create_locations_xlsx_file(self, data_rows, headers=None):
        if headers is None:
            headers = ["Partner information ", "Location Information "]

        stream = io.BytesIO()
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        sheet.append(headers)
        sheet.append(["Help text: ip_numbers is a JSON array of vendor numbers. Ex: [\"VN001\", \"VN002\"]"])

        for row_data in data_rows:
            sheet.append(row_data)

        workbook.save(stream)
        stream.seek(0)
        return stream

    def test_import_locations_successful(self):
        self.assertEqual(PointOfInterest.objects.count(), 1)
        data_rows = [
            [json.dumps(["IP_VN_001"]), "New Clinic A", "Health Facility", 43.1, 25.1, "PCODE_A"],
            [json.dumps(["IP_VN_001", "IP_VN_002"]), "New School B", "School", 43.2, 25.2, "PCODE_B"],
            [json.dumps([]), "Warehouse C", "Other", 43.3, 25.3, "PCODE_C"],
            ["", "Distribution Point D", "Other", 43.4, 25.4, "PCODE_D"],
        ]
        xlsx_file = self._create_locations_xlsx_file(data_rows)
        xlsx_file.name = "success_locations_import.xlsx"

        response = self.forced_auth_req(
            'post',
            self.import_url,
            user=self.admin_user,
            data={'file': xlsx_file},
            request_format='multipart'
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data, {"valid": True})
        self.assertEqual(PointOfInterest.objects.count(), 1 + 4)

        loc_a = PointOfInterest.objects.get(p_code="PCODE_A")
        self.assertEqual(loc_a.name, "New Clinic A")
        self.assertEqual(loc_a.poi_type, self.poi_type_health)
        self.assertEqual(loc_a.point.y, 43.1)  # Latitude
        self.assertEqual(loc_a.point.x, 25.1)  # Longitude
        self.assertFalse(loc_a.is_active)
        self.assertEqual(loc_a.created_by, self.admin_user)
        self.assertIn(self.partner_org1, loc_a.partner_organizations.all())
        self.assertEqual(loc_a.partner_organizations.count(), 1)

        loc_b = PointOfInterest.objects.get(p_code="PCODE_B")
        self.assertEqual(loc_b.name, "New School B")
        self.assertEqual(loc_b.poi_type, self.poi_type_school)
        self.assertIn(self.partner_org1, loc_b.partner_organizations.all())
        self.assertIn(self.partner_org2, loc_b.partner_organizations.all())
        self.assertEqual(loc_b.partner_organizations.count(), 2)

        loc_c = PointOfInterest.objects.get(p_code="PCODE_C")
        self.assertEqual(loc_c.partner_organizations.count(), 0)
        loc_d = PointOfInterest.objects.get(p_code="PCODE_D")
        self.assertEqual(loc_d.partner_organizations.count(), 0)

    def test_import_locations_no_file_provided(self):
        response = self.forced_auth_req(
            'post',
            self.import_url,
            user=self.admin_user,
            data={},
            request_format='multipart'
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('file', response.data)
        self.assertEqual(response.data['file'][0].code, 'required')

    def test_import_locations_with_some_invalid_rows_returns_error_file(self):
        initial_poi_count = PointOfInterest.objects.count()
        data_rows = [
            [json.dumps(["IP_VN_001"]), "Valid Location 1", "School", 5.5, 6.6, "PCODE_VALID1"],  # Valid
            ["", "Another Name", "School", 7.7, 8.8, "PCODE_EXISTING"],  # Invalid: Duplicate P-Code
            ["", "Existing General Hospital", "Health Facility", 1.1, 2.2, "PCODE_DUPE_NAME"],  # Invalid: Duplicate Name
            ["", "Bad Type Loc", "NonExistentType", 3.3, 4.4, "PCODE_BAD_TYPE"],  # Invalid: POI Type
            ["", "Bad Coords", "School", "not-a-latitude", 5.5, "PCODE_BAD_COORDS"],  # Invalid: Latitude
            ["not-a-json", "Bad IP JSON", "School", 6.6, 7.7, "PCODE_BAD_JSON"],  # Invalid: IP Numbers JSON format
            [json.dumps(["NON_EXISTENT_VN"]), "Bad IP", "Other", 8.8, 9.9, "PCODE_BAD_IP"],  # Invalid: Vendor number
        ]
        xlsx_file = self._create_locations_xlsx_file(data_rows)
        xlsx_file.name = "partial_fail_locations_import.xlsx"

        response = self.forced_auth_req(
            'post',
            self.import_url,
            user=self.admin_user,
            data={'file': xlsx_file},
            request_format='multipart'
        )

        self.assertIsInstance(response, HttpResponse)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response['Content-Disposition'].startswith('attachment; filename="checked_partial_fail_locations_import.xlsx"'))
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        self.assertEqual(PointOfInterest.objects.count(), initial_poi_count + 1)
        self.assertTrue(PointOfInterest.objects.filter(p_code="PCODE_VALID1").exists())
        self.assertFalse(PointOfInterest.objects.filter(p_code__in=["PCODE_DUPE_NAME", "PCODE_BAD_TYPE"]).exists())

        returned_file_content = io.BytesIO(response.content)
        workbook = openpyxl.load_workbook(returned_file_content)
        sheet = workbook.active
        errors_col_idx = sheet.max_column

        self.assertEqual(sheet.cell(row=3, column=errors_col_idx).value, "Success")
        self.assertIn("Point of interest already exists", sheet.cell(row=4, column=errors_col_idx).value)
        self.assertIn("Point of interest already exists", sheet.cell(row=5, column=errors_col_idx).value)
        self.assertIn("Point of interest type does not exist", sheet.cell(row=6, column=errors_col_idx).value)
        self.assertIn("Invalid latitude", sheet.cell(row=7, column=errors_col_idx).value)
        self.assertIn("Invalid 'IP Numbers' format. Must be a valid JSON list.", sheet.cell(row=8, column=errors_col_idx).value)
        self.assertIn("Object with organization__vendor_number=NON_EXISTENT_VN does not exist", sheet.cell(row=9, column=errors_col_idx).value)

    def test_import_empty_file_is_valid(self):
        initial_poi_count = PointOfInterest.objects.count()
        xlsx_file = self._create_locations_xlsx_file([])
        xlsx_file.name = "empty_import.xlsx"

        response = self.forced_auth_req(
            'post',
            self.import_url,
            user=self.admin_user,
            data={'file': xlsx_file},
            request_format='multipart'
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data, {"valid": True})
        self.assertEqual(PointOfInterest.objects.count(), initial_poi_count)

    def test_import_with_malformed_file_not_xlsx(self):
        """
        Ensure that uploading a file that is not a valid XLSX format is handled gracefully.
        """
        malformed_file = io.BytesIO(b"this is just a text file, not a spreadsheet.")
        malformed_file.name = "not_an_excel.txt"

        with patch(
            'etools.applications.last_mile.admin_panel.views.CsvImporter.import_locations',
            side_effect=InvalidFileException("File is not a valid zip file")
        ) as mock_importer:

            with self.assertRaises(InvalidFileException):
                self.forced_auth_req(
                    'post',
                    self.import_url,
                    user=self.admin_user,
                    data={'file': malformed_file},
                    request_format='multipart'
                )
            mock_importer.assert_called_once()

    def test_import_with_duplicate_p_code_within_same_file(self):
        """
        Tests that if the same p_code is used for two rows in one file,
        the first is created and the second fails validation.
        """
        initial_poi_count = PointOfInterest.objects.count()
        data_rows = [
            ["", "First with Dupe P-Code", "School", 10.1, 10.2, "PCODE_INTERNAL_DUPE"],
            ["", "Second with Dupe P-Code", "School", 11.1, 11.2, "PCODE_INTERNAL_DUPE"],
        ]
        xlsx_file = self._create_locations_xlsx_file(data_rows)
        xlsx_file.name = "internal_dupe_pcode.xlsx"

        response = self.forced_auth_req(
            'post',
            self.import_url,
            user=self.admin_user,
            data={'file': xlsx_file},
            request_format='multipart'
        )

        self.assertIsInstance(response, HttpResponse)
        self.assertEqual(response.status_code, 200)

        self.assertEqual(PointOfInterest.objects.count(), initial_poi_count + 1)
        self.assertTrue(PointOfInterest.objects.filter(p_code="PCODE_INTERNAL_DUPE").exists())

        returned_file_content = io.BytesIO(response.content)
        workbook = openpyxl.load_workbook(returned_file_content)
        sheet = workbook.active
        errors_col_idx = sheet.max_column

        self.assertEqual(sheet.cell(row=3, column=errors_col_idx).value, "Success")
        self.assertIn("Point of interest already exists", sheet.cell(row=4, column=errors_col_idx).value)

    def test_import_with_missing_required_fields(self):
        """
        Test that rows with missing required data (e.g., empty p_code_location)
        are caught by the serializer's default validation.
        """
        data_rows = [
            ["", "", "", 20.1, 20.2, ""],
        ]
        xlsx_file = self._create_locations_xlsx_file(data_rows)
        xlsx_file.name = "missing_required.xlsx"

        response = self.forced_auth_req(
            'post',
            self.import_url,
            user=self.admin_user,
            data={'file': xlsx_file},
            request_format='multipart'
        )

        self.assertIsInstance(response, HttpResponse)
        self.assertEqual(response.status_code, 200)

        returned_file_content = io.BytesIO(response.content)
        workbook = openpyxl.load_workbook(returned_file_content)
        sheet = workbook.active
        errors_col_idx = sheet.max_column
        error_message = sheet.cell(row=3, column=errors_col_idx).value
        self.assertIn("{'p_code_location': [ErrorDetail(string='This field may not be null.'", error_message)

    @patch('etools.applications.last_mile.models.PointOfInterest.objects.create')
    def test_import_handles_unexpected_database_error_during_creation(self, mock_poi_create):
        from django.db import IntegrityError

        initial_poi_count = PointOfInterest.objects.count()
        mock_poi_create.side_effect = IntegrityError("Forced database error for testing.")

        data_rows = [
            ["", "Location that will fail", "School", 30.1, 30.2, "PCODE_DB_FAIL"],
        ]
        xlsx_file = self._create_locations_xlsx_file(data_rows)
        xlsx_file.name = "db_error.xlsx"

        response = self.forced_auth_req(
            'post',
            self.import_url,
            user=self.admin_user,
            data={'file': xlsx_file},
            request_format='multipart'
        )

        self.assertIsInstance(response, HttpResponse)
        self.assertEqual(response.status_code, 200)

        self.assertEqual(PointOfInterest.objects.count(), initial_poi_count)

        returned_file_content = io.BytesIO(response.content)
        workbook = openpyxl.load_workbook(returned_file_content)
        sheet = workbook.active
        errors_col_idx = sheet.max_column
        error_message = sheet.cell(row=3, column=errors_col_idx).value

        self.assertEqual(error_message, "Forced database error for testing.")

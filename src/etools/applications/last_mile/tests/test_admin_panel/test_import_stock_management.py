import io
from datetime import datetime
from unittest import mock

import openpyxl
from rest_framework import status
from rest_framework.reverse import reverse

from etools.applications.core.tests.cases import BaseTenantTestCase
from etools.applications.last_mile.admin_panel.constants import (
    ADMIN_PANEL_APP_NAME,
    STOCK_MANAGEMENT_ADMIN_PANEL,
    STOCK_MANAGEMENT_ADMIN_PANEL_PERMISSION,
)
from etools.applications.last_mile.models import Item, Transfer
from etools.applications.last_mile.tests.factories import MaterialFactory, PointOfInterestFactory
from etools.applications.organizations.tests.factories import OrganizationFactory
from etools.applications.partners.tests.factories import PartnerFactory
from etools.applications.users.tests.factories import UserPermissionFactory


class TestStockAdminViewSetImport(BaseTenantTestCase):

    @classmethod
    def setUpTestData(cls):
        cls.admin_user = UserPermissionFactory(
            realms__data=['LMSM Admin Panel', 'IP LM Editor'],
            username="stock_importer_admin",
            email="stock_importer_admin@example.com",
            is_staff=True,
            perms=[STOCK_MANAGEMENT_ADMIN_PANEL_PERMISSION]
        )

        cls.org1 = OrganizationFactory(name="Stock Org 1", vendor_number="VN_STOCK_001")
        cls.partner1 = PartnerFactory(organization=cls.org1)

        cls.material1 = MaterialFactory(number="MAT001")
        cls.material2 = MaterialFactory(number="MAT002")

        cls.poi1 = PointOfInterestFactory(name="Stock POI 1", p_code="PCODE_STOCK_001", partner_organizations=[cls.partner1])
        cls.unicef_warehouse = PointOfInterestFactory(name="UNICEF Warehouse")

        cls.import_url = reverse(f'{ADMIN_PANEL_APP_NAME}:{STOCK_MANAGEMENT_ADMIN_PANEL}--import-file')

        cls.unauthorized_user = UserPermissionFactory(
            username="unauthorized_stock_user",
            email="unauthorized_stock_user@example.com",
            is_staff=True,
        )
        cls.org2 = OrganizationFactory(name="Stock Org 2", vendor_number="VN_STOCK_002")
        cls.partner2 = PartnerFactory(organization=cls.org2)
        cls.poi2_for_partner2 = PointOfInterestFactory(name="Stock POI 2", p_code="PCODE_STOCK_002", partner_organizations=[cls.partner2])

    def _create_xlsx_file(self, data_rows, headers=None):
        if headers is None:
            headers = ["Partner information ", "Stock level information ", None, None, None, None, "Location Information "]

        stream = io.BytesIO()
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        sheet.append(headers)
        sheet.append(["Help text / Description row (ignored by importer)"])

        for row_data in data_rows:
            sheet.append(row_data)

        workbook.save(stream)
        stream.seek(0)
        return stream

    @mock.patch('etools.applications.last_mile.models.PointOfInterest.objects.get_unicef_warehouses')
    def test_import_stock_successful(self, mock_get_unicef_warehouses):
        mock_get_unicef_warehouses.return_value = self.unicef_warehouse
        self.assertEqual(Transfer.all_objects.count(), 0)
        self.assertEqual(Item.objects.count(), 0)

        expiration_date = datetime(2099, 12, 31, 23, 59, 59)
        data_rows = [
            ["VN_STOCK_001", "MAT001", 100, "BOX", expiration_date, "BATCH001", "PCODE_STOCK_001"],
            ["VN_STOCK_001", "MAT002", 50, "EA", expiration_date, "", "PCODE_STOCK_001"],
        ]
        xlsx_file = self._create_xlsx_file(data_rows)
        xlsx_file.name = "success_stock_import.xlsx"

        response = self.forced_auth_req(
            'post',
            self.import_url,
            user=self.admin_user,
            data={'file': xlsx_file},
            request_format='multipart'
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data, {"valid": True})

        self.assertEqual(Transfer.all_objects.count(), 2)
        self.assertEqual(Item.objects.count(), 2)

        transfer1 = Transfer.all_objects.get(items__material=self.material1)
        self.assertEqual(transfer1.partner_organization, self.partner1)
        self.assertEqual(transfer1.destination_point, self.poi1)
        self.assertEqual(transfer1.origin_point, self.unicef_warehouse)
        self.assertEqual(transfer1.transfer_type, Transfer.DELIVERY)
        self.assertEqual(transfer1.approval_status, "PENDING")

        item1 = transfer1.items.first()
        self.assertEqual(item1.material, self.material1)
        self.assertEqual(item1.quantity, 100)
        self.assertEqual(item1.uom, "BOX")
        self.assertEqual(item1.batch_id, "BATCH001")
        self.assertEqual(item1.expiry_date.year, expiration_date.year)

        item2 = Item.objects.get(material=self.material2)
        self.assertEqual(item2.quantity, 50)
        self.assertEqual(item2.uom, "EA")
        self.assertEqual(item2.batch_id, None)

    def test_import_stock_no_file_provided(self):
        response = self.forced_auth_req(
            'post',
            self.import_url,
            user=self.admin_user,
            data={},
            request_format='multipart'
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('file', response.data)

    @mock.patch('etools.applications.last_mile.models.PointOfInterest.objects.get_unicef_warehouses')
    def test_import_stock_with_some_invalid_rows(self, mock_get_unicef_warehouses):
        mock_get_unicef_warehouses.return_value = self.unicef_warehouse
        initial_transfer_count = Transfer.objects.count()

        expiration_date = datetime(2099, 12, 31, 23, 59, 59)
        data_rows = [
            ["VN_STOCK_001", "MAT001", 100, "BOX", expiration_date, "BATCH001", "PCODE_STOCK_001"],  # Valid
            ["VN_STOCK_999", "MAT001", 10, "BOX", expiration_date, "B002", "PCODE_STOCK_001"],    # Invalid ip_number
            ["VN_STOCK_001", "MAT999", 20, "BOX", expiration_date, "B003", "PCODE_STOCK_001"],    # Invalid material
            ["VN_STOCK_001", "MAT001", -50, "BOX", expiration_date, "B004", "PCODE_STOCK_001"],    # Invalid quantity
            ["VN_STOCK_001", "MAT001", 30, "BOX", expiration_date, "B005", "PCODE_STOCK_999"],    # Invalid p_code
            ["VN_STOCK_001", "MAT001", 40, "BOX", "not-a-date", "B006", "PCODE_STOCK_001"],      # Invalid date
        ]
        xlsx_file = self._create_xlsx_file(data_rows)
        xlsx_file.name = "partial_fail_stock_import.xlsx"

        response = self.forced_auth_req(
            'post',
            self.import_url,
            user=self.admin_user,
            data={'file': xlsx_file},
            request_format='multipart'
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response['Content-Disposition'].startswith('attachment; filename="checked_partial_fail_stock_import.xlsx"'))
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        self.assertEqual(Transfer.all_objects.count(), initial_transfer_count + 1)
        self.assertTrue(Item.objects.filter(batch_id="BATCH001").exists())

        returned_file_content = io.BytesIO(response.content)
        workbook = openpyxl.load_workbook(returned_file_content)
        sheet = workbook.active
        errors_col_idx = sheet.max_column

        self.assertEqual(sheet.cell(row=3, column=errors_col_idx).value, "Success")
        self.assertIn("Partner Organization not found", sheet.cell(row=4, column=errors_col_idx).value)
        self.assertIn("Material not found", sheet.cell(row=5, column=errors_col_idx).value)
        self.assertIn("invalid_quantity", sheet.cell(row=6, column=errors_col_idx).value)
        self.assertIn("Point of interest not found", sheet.cell(row=7, column=errors_col_idx).value)
        self.assertIn("Datetime has wrong format. Use one of these formats instead: YYYY-MM-DDThh:mm", sheet.cell(row=8, column=errors_col_idx).value)

    def test_import_stock_all_rows_invalid_returns_error_file(self):
        initial_transfer_count = Transfer.objects.count()
        expiration_date = datetime(2099, 12, 31, 23, 59, 59)
        data_rows = [
            ["VN_STOCK_999", "MAT001", 10, "BOX", expiration_date, "B002", "PCODE_STOCK_001"],
            ["VN_STOCK_001", "MAT999", 20, "BOX", expiration_date, "B003", "PCODE_STOCK_001"],
        ]
        xlsx_file = self._create_xlsx_file(data_rows)
        xlsx_file.name = "all_fail_stock_import.xlsx"

        response = self.forced_auth_req(
            'post',
            self.import_url,
            user=self.admin_user,
            data={'file': xlsx_file},
            request_format='multipart'
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response['Content-Disposition'].startswith('attachment; filename="checked_all_fail_stock_import.xlsx"'))

        self.assertEqual(Transfer.objects.count(), initial_transfer_count)

    def test_import_unauthorized_user_fails(self):
        expiration_date = datetime(2099, 12, 31)
        data_rows = [
            ["VN_STOCK_001", "MAT001", 100, "BOX", expiration_date, "BATCH_A", "PCODE_STOCK_001"],
        ]
        xlsx_file = self._create_xlsx_file(data_rows)
        xlsx_file.name = "unauthorized_import.xlsx"

        response = self.forced_auth_req(
            'post',
            self.import_url,
            user=self.unauthorized_user,
            data={'file': xlsx_file},
            request_format='multipart'
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(Transfer.objects.count(), 0)

    def test_import_with_missing_required_fields_fails(self):
        expiration_date = datetime(2099, 12, 31)
        data_rows = [
            ["", "MAT001", 100, "BOX", expiration_date, "BATCH_A", "PCODE_STOCK_001"],        # Missing ip_number
            ["VN_STOCK_001", "", 50, "EA", expiration_date, "BATCH_B", "PCODE_STOCK_001"],     # Missing material_number
            ["VN_STOCK_001", "MAT001", "", "BOX", expiration_date, "BATCH_C", "PCODE_STOCK_001"],  # Missing quantity
        ]
        xlsx_file = self._create_xlsx_file(data_rows)
        xlsx_file.name = "missing_fields.xlsx"

        response = self.forced_auth_req('post', self.import_url, user=self.admin_user, data={'file': xlsx_file}, request_format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(Transfer.objects.count(), 0)

        returned_file = io.BytesIO(response.content)
        workbook = openpyxl.load_workbook(returned_file)
        sheet = workbook.active
        errors_col_idx = sheet.max_column

        self.assertIn("This field may not be null", sheet.cell(row=3, column=errors_col_idx).value)
        self.assertIn("This field may not be null", sheet.cell(row=4, column=errors_col_idx).value)
        self.assertIn("This field may not be null", sheet.cell(row=5, column=errors_col_idx).value)

    def test_import_with_zero_quantity_fails(self):
        expiration_date = datetime(2099, 12, 31)
        data_rows = [
            ["VN_STOCK_001", "MAT001", 0, "BOX", expiration_date, "BATCH_ZERO", "PCODE_STOCK_001"],
        ]
        xlsx_file = self._create_xlsx_file(data_rows)
        xlsx_file.name = "zero_quantity.xlsx"

        response = self.forced_auth_req('post', self.import_url, user=self.admin_user, data={'file': xlsx_file}, request_format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(Transfer.objects.count(), 0)
        returned_file = io.BytesIO(response.content)
        workbook = openpyxl.load_workbook(returned_file)
        sheet = workbook.active
        self.assertIn("invalid_quantity", sheet.cell(row=3, column=sheet.max_column).value)

    def test_import_empty_file_succeeds_with_no_changes(self):
        initial_transfer_count = Transfer.objects.count()
        xlsx_file = self._create_xlsx_file(data_rows=[])  # No data rows
        xlsx_file.name = "empty_import.xlsx"

        response = self.forced_auth_req('post', self.import_url, user=self.admin_user, data={'file': xlsx_file}, request_format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data, {"valid": True})
        self.assertEqual(Transfer.objects.count(), initial_transfer_count)

    @mock.patch('etools.applications.last_mile.models.PointOfInterest.objects.get_unicef_warehouses')
    def test_import_with_extra_columns_is_ignored_and_succeeds(self, mock_get_unicef_warehouses):
        mock_get_unicef_warehouses.return_value = self.unicef_warehouse
        expiration_date = datetime(2099, 12, 31)
        headers = ["Partner information ", "Stock level information ", None, None, None, None, "Location Information "]
        data_rows = [
            ["VN_STOCK_001", "MAT001", 100, "BOX", expiration_date, "BATCH_EXTRA", "PCODE_STOCK_001", "This is an extra note."],
        ]
        xlsx_file = self._create_xlsx_file(data_rows, headers=headers)
        xlsx_file.name = "extra_columns.xlsx"

        response = self.forced_auth_req('post', self.import_url, user=self.admin_user, data={'file': xlsx_file}, request_format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data, {"valid": True})
        self.assertEqual(Transfer.all_objects.count(), 1)
        self.assertEqual(Transfer.all_objects.first().approval_status, "PENDING")
        self.assertTrue(Item.objects.filter(batch_id="BATCH_EXTRA").exists())

    @mock.patch('etools.applications.last_mile.models.PointOfInterest.objects.get_unicef_warehouses')
    def test_import_pcode_not_linked_to_partner_succeeds(self, mock_get_unicef_warehouses):
        mock_get_unicef_warehouses.return_value = self.unicef_warehouse
        expiration_date = datetime(2099, 12, 31)
        data_rows = [
            ["VN_STOCK_001", "MAT001", 10, "EA", expiration_date, "BATCH_CROSS", "PCODE_STOCK_002"],
        ]
        xlsx_file = self._create_xlsx_file(data_rows)
        xlsx_file.name = "cross_partner_poi.xlsx"

        response = self.forced_auth_req('post', self.import_url, user=self.admin_user, data={'file': xlsx_file}, request_format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data, {"valid": True})

        self.assertEqual(Transfer.all_objects.count(), 1)
        transfer = Transfer.all_objects.first()
        self.assertEqual(transfer.partner_organization, self.partner1)
        self.assertEqual(transfer.destination_point, self.poi2_for_partner2)
        self.assertEqual(transfer.approval_status, "PENDING")

    @mock.patch('etools.applications.last_mile.models.PointOfInterest.objects.get_unicef_warehouses')
    def test_imported_transfer_has_pending_approval_status(self, mock_get_unicef_warehouses):
        mock_get_unicef_warehouses.return_value = self.unicef_warehouse
        expiration_date = datetime(2099, 12, 31)
        data_rows = [
            ["VN_STOCK_001", "MAT001", 100, "BOX", expiration_date, "BATCH_APPROVAL", "PCODE_STOCK_001"],
        ]
        xlsx_file = self._create_xlsx_file(data_rows)
        xlsx_file.name = "approval_status_test.xlsx"

        response = self.forced_auth_req('post', self.import_url, user=self.admin_user, data={'file': xlsx_file}, request_format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data, {"valid": True})

        transfer = Transfer.all_objects.first()
        self.assertIsNotNone(transfer)
        self.assertEqual(transfer.approval_status, Transfer.ApprovalStatus.PENDING)

    @mock.patch('etools.applications.last_mile.models.PointOfInterest.objects.get_unicef_warehouses')
    def test_imported_pending_transfer_not_in_default_queryset(self, mock_get_unicef_warehouses):
        mock_get_unicef_warehouses.return_value = self.unicef_warehouse
        expiration_date = datetime(2099, 12, 31)
        data_rows = [
            ["VN_STOCK_001", "MAT001", 50, "EA", expiration_date, "BATCH_PENDING", "PCODE_STOCK_001"],
        ]
        xlsx_file = self._create_xlsx_file(data_rows)
        xlsx_file.name = "pending_queryset_test.xlsx"

        response = self.forced_auth_req('post', self.import_url, user=self.admin_user, data={'file': xlsx_file}, request_format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data, {"valid": True})

        self.assertEqual(Transfer.all_objects.count(), 1)
        self.assertEqual(Transfer.objects.count(), 0)

    @mock.patch('etools.applications.last_mile.models.PointOfInterest.objects.get_unicef_warehouses')
    def test_multiple_imports_all_have_pending_status(self, mock_get_unicef_warehouses):
        mock_get_unicef_warehouses.return_value = self.unicef_warehouse
        expiration_date = datetime(2099, 12, 31)
        data_rows = [
            ["VN_STOCK_001", "MAT001", 10, "BOX", expiration_date, "BATCH_A", "PCODE_STOCK_001"],
            ["VN_STOCK_001", "MAT002", 20, "EA", expiration_date, "BATCH_B", "PCODE_STOCK_001"],
            ["VN_STOCK_002", "MAT001", 30, "BOX", expiration_date, "BATCH_C", "PCODE_STOCK_002"],
        ]
        xlsx_file = self._create_xlsx_file(data_rows)
        xlsx_file.name = "multiple_pending_imports.xlsx"

        response = self.forced_auth_req('post', self.import_url, user=self.admin_user, data={'file': xlsx_file}, request_format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data, {"valid": True})

        self.assertEqual(Transfer.all_objects.count(), 3)
        self.assertEqual(Transfer.objects.count(), 0)

        for transfer in Transfer.all_objects.all():
            self.assertEqual(transfer.approval_status, Transfer.ApprovalStatus.PENDING)

    @mock.patch('etools.applications.last_mile.models.PointOfInterest.objects.get_unicef_warehouses')
    def test_imported_items_linked_to_pending_transfers(self, mock_get_unicef_warehouses):
        mock_get_unicef_warehouses.return_value = self.unicef_warehouse
        expiration_date = datetime(2099, 12, 31)
        data_rows = [
            ["VN_STOCK_001", "MAT001", 100, "BOX", expiration_date, "BATCH_LINK", "PCODE_STOCK_001"],
        ]
        xlsx_file = self._create_xlsx_file(data_rows)
        xlsx_file.name = "item_link_test.xlsx"

        response = self.forced_auth_req('post', self.import_url, user=self.admin_user, data={'file': xlsx_file}, request_format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        transfer = Transfer.all_objects.first()
        self.assertIsNotNone(transfer)
        self.assertEqual(transfer.approval_status, Transfer.ApprovalStatus.PENDING)

        item = Item.objects.first()
        self.assertIsNotNone(item)
        self.assertEqual(item.transfer, transfer)
        self.assertEqual(item.batch_id, "BATCH_LINK")
        self.assertEqual(item.quantity, 100)

    @mock.patch('etools.applications.last_mile.models.PointOfInterest.objects.get_unicef_warehouses')
    def test_import_with_validation_error_no_pending_transfer_created(self, mock_get_unicef_warehouses):
        mock_get_unicef_warehouses.return_value = self.unicef_warehouse
        expiration_date = datetime(2099, 12, 31)
        data_rows = [
            ["VN_STOCK_999", "MAT001", 10, "BOX", expiration_date, "BATCH_ERR1", "PCODE_STOCK_001"],  # Invalid vendor
            ["VN_STOCK_001", "MAT999", 20, "BOX", expiration_date, "BATCH_ERR2", "PCODE_STOCK_001"],  # Invalid material
        ]
        xlsx_file = self._create_xlsx_file(data_rows)
        xlsx_file.name = "validation_error_test.xlsx"

        initial_count = Transfer.all_objects.count()

        response = self.forced_auth_req('post', self.import_url, user=self.admin_user, data={'file': xlsx_file}, request_format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        self.assertEqual(Transfer.all_objects.count(), initial_count)

    @mock.patch('etools.applications.last_mile.models.PointOfInterest.objects.get_unicef_warehouses')
    def test_import_creates_transfer_with_correct_transfer_type(self, mock_get_unicef_warehouses):
        mock_get_unicef_warehouses.return_value = self.unicef_warehouse
        expiration_date = datetime(2099, 12, 31)
        data_rows = [
            ["VN_STOCK_001", "MAT001", 100, "BOX", expiration_date, "BATCH_TYPE", "PCODE_STOCK_001"],
        ]
        xlsx_file = self._create_xlsx_file(data_rows)
        xlsx_file.name = "transfer_type_test.xlsx"

        response = self.forced_auth_req('post', self.import_url, user=self.admin_user, data={'file': xlsx_file}, request_format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        transfer = Transfer.all_objects.first()
        self.assertIsNotNone(transfer)
        self.assertEqual(transfer.transfer_type, Transfer.DELIVERY)
        self.assertEqual(transfer.approval_status, Transfer.ApprovalStatus.PENDING)

    @mock.patch('etools.applications.last_mile.models.PointOfInterest.objects.get_unicef_warehouses')
    def test_import_sets_origin_and_destination_correctly(self, mock_get_unicef_warehouses):
        mock_get_unicef_warehouses.return_value = self.unicef_warehouse
        expiration_date = datetime(2099, 12, 31)
        data_rows = [
            ["VN_STOCK_001", "MAT001", 50, "EA", expiration_date, "BATCH_ORIGIN", "PCODE_STOCK_001"],
        ]
        xlsx_file = self._create_xlsx_file(data_rows)
        xlsx_file.name = "origin_destination_test.xlsx"

        response = self.forced_auth_req('post', self.import_url, user=self.admin_user, data={'file': xlsx_file}, request_format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        transfer = Transfer.all_objects.first()
        self.assertIsNotNone(transfer)
        self.assertEqual(transfer.origin_point, self.unicef_warehouse)
        self.assertEqual(transfer.destination_point, self.poi1)
        self.assertEqual(transfer.approval_status, Transfer.ApprovalStatus.PENDING)

    @mock.patch('etools.applications.last_mile.models.PointOfInterest.objects.get_unicef_warehouses')
    def test_import_partial_success_only_valid_rows_create_pending_transfers(self, mock_get_unicef_warehouses):
        mock_get_unicef_warehouses.return_value = self.unicef_warehouse
        expiration_date = datetime(2099, 12, 31)
        data_rows = [
            ["VN_STOCK_001", "MAT001", 100, "BOX", expiration_date, "BATCH_VALID1", "PCODE_STOCK_001"],  # Valid
            ["VN_STOCK_999", "MAT001", 10, "BOX", expiration_date, "BATCH_INVALID", "PCODE_STOCK_001"],  # Invalid vendor
            ["VN_STOCK_001", "MAT002", 50, "EA", expiration_date, "BATCH_VALID2", "PCODE_STOCK_001"],   # Valid
        ]
        xlsx_file = self._create_xlsx_file(data_rows)
        xlsx_file.name = "partial_success_approval_test.xlsx"

        response = self.forced_auth_req('post', self.import_url, user=self.admin_user, data={'file': xlsx_file}, request_format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        self.assertEqual(Transfer.all_objects.count(), 2)
        for transfer in Transfer.all_objects.all():
            self.assertEqual(transfer.approval_status, Transfer.ApprovalStatus.PENDING)

        self.assertEqual(Item.objects.count(), 2)
        self.assertTrue(Item.objects.filter(batch_id="BATCH_VALID1").exists())
        self.assertTrue(Item.objects.filter(batch_id="BATCH_VALID2").exists())
        self.assertFalse(Item.objects.filter(batch_id="BATCH_INVALID").exists())

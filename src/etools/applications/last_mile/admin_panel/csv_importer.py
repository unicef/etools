import io
import json
from functools import partial

from django.db import transaction

import openpyxl

from etools.applications.last_mile.admin_panel.serializers import (
    LocationImportSerializer,
    StockManagementImportSerializer,
    UserImportSerializer,
)


class CsvReportHandler:
    def __init__(self, file_bytes):
        stream = io.BytesIO(file_bytes)
        self.wb = openpyxl.load_workbook(stream)
        self.ws = self.wb.active
        self.status_col = self.ws.max_column + 1
        self.ws.cell(row=1, column=self.status_col, value='Import Status')

    def get_rows(self):
        for index, row in enumerate(self.ws.iter_rows(min_row=3, values_only=True), start=3):
            if not any(row):
                continue
            yield index, row

    def validate_header_row(self, file_for=""):
        expected_headers = {
            "users": ["Partner information ", "User Information ", None, None, "Location Information"],
            "locations": ["Partner information ", "Location Information "],
            "stock": ["Partner Information ", "Stock Level Information ", None, None, None, None, "Location Information (This information can be found in the Admin Module under Manage Locations) "],
        }

        if file_for not in expected_headers:
            return False
        row_num = 1
        actual_headers = [
            self.ws.cell(row=row_num, column=col).value
            for col in range(1, len(expected_headers[file_for]) + 1)
        ]

        return actual_headers == expected_headers[file_for]

    def write_status(self, row_index, message):
        self.ws.cell(row=row_index, column=self.status_col, value=message)

    def get_processed_file(self):
        out = io.BytesIO()
        self.wb.save(out)
        out.seek(0)
        return out


class CsvImporter:
    def _process_user_row(self, row_data, country_schema, user):
        ip_number, first_name, last_name, email, poi_str, *_ = row_data

        try:
            point_of_interests = json.loads(poi_str) if poi_str else []
        except (ValueError, TypeError):
            return False, "Invalid 'Point of Interests' format. Must be a valid JSON list."

        data = {
            'ip_number': ip_number,
            'first_name': first_name,
            'last_name': last_name,
            'email': email,
            'point_of_interests': point_of_interests,
        }
        serializer = UserImportSerializer(data=data)

        if not serializer.is_valid():
            return False, str(serializer.errors)

        created, object_data = serializer.create(serializer.validated_data, country_schema, user)
        if not created:
            return False, object_data

        return True, "Success"

    def _process_location_row(self, row_data, user):
        ip_numbers_str, location_name, primary_type_name, latitude, longitude, p_code_location, *_ = row_data

        try:
            ip_numbers = json.loads(ip_numbers_str) if ip_numbers_str else []
        except (ValueError, TypeError):
            return False, "Invalid 'IP Numbers' format. Must be a valid JSON list."

        data = {
            'ip_numbers': ip_numbers,
            'location_name': location_name,
            'primary_type_name': primary_type_name,
            'latitude': latitude,
            'longitude': longitude,
            'p_code_location': p_code_location,
        }
        serializer = LocationImportSerializer(data=data)

        if not serializer.is_valid():
            return False, str(serializer.errors)

        created, object_data = serializer.create(serializer.validated_data, user)
        if not created:
            return False, object_data

        return True, "Success"

    def _process_stock_row(self, row_data, user):
        ip_number, material_number, quantity, uom, expiration_date, batch_id, p_code, *_ = row_data

        data = {
            'ip_number': ip_number,
            'material_number': material_number,
            'quantity': quantity,
            'uom': uom,
            'expiration_date': expiration_date,
            'batch_id': batch_id,
            'p_code': p_code
        }

        serializer = StockManagementImportSerializer(data=data)

        if not serializer.is_valid():
            return False, str(serializer.errors)
        try:
            created, object_data = serializer.create(serializer.validated_data, user)
        except Exception as ex:
            return False, str(ex)
        if not created:
            return False, object_data

        return True, "Success"

    def _perform_import(self, file, row_processor, file_for):
        file_handler = CsvReportHandler(file.read())
        is_valid_file = file_handler.validate_header_row(file_for=file_for)
        if not is_valid_file:
            processed_file = file_handler.get_processed_file()
            return False, processed_file
        import_is_fully_successful = True

        with transaction.atomic():
            for index, row_data in file_handler.get_rows():
                is_valid, status_message = row_processor(row_data)

                if not is_valid:
                    import_is_fully_successful = False

                file_handler.write_status(index, status_message)

        processed_file = file_handler.get_processed_file()
        return import_is_fully_successful, processed_file

    def import_users(self, file, country_schema, user):
        processor = partial(self._process_user_row, country_schema=country_schema, user=user)
        return self._perform_import(file, processor, "users")

    def import_locations(self, file, user):
        processor = partial(self._process_location_row, user=user)
        return self._perform_import(file, processor, "locations")

    def import_stock(self, file, user):
        processor = partial(self._process_stock_row, user=user)
        return self._perform_import(file, processor, 'stock')
